#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
报销单模板匹配与自适应学习引擎
企业自主提供模板，Skill自动学习映射关系
"""

import sys
import json
import re
import argparse
from pathlib import Path
from datetime import datetime

# 尝试导入yaml和openpyxl
try:
    import yaml
except ImportError:
    yaml = None

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl未安装。请运行: pip install openpyxl")
    sys.exit(1)


class TemplateMatcher:
    """模板匹配器"""

    def __init__(self, template_path=None, config_path=None):
        """
        初始化模板匹配器

        Args:
            template_path: 报销单模板路径（.xlsx）
            config_path: 配置文件路径（可选，保存映射关系）
        """
        self.template_path = template_path
        self.config_path = config_path
        self.mapping = {}
        self.config = {}
        self.cache = {}

        # 加载配置
        if config_path and Path(config_path).exists():
            self._load_config(config_path)

        # 加载缓存的映射
        cache_path = Path(config_path).parent / 'template_cache.yaml' if config_path else None
        if cache_path and cache_path.exists():
            self._load_cache(cache_path)

    def _load_config(self, path):
        """加载配置文件"""
        with open(path, 'r', encoding='utf-8') as f:
            if yaml and path.endswith(('.yaml', '.yml')):
                self.config = yaml.safe_load(f)
            else:
                self.config = json.load(f)

        self.mapping = self.config.get('template', {}).get('field_mapping', {})

    def _load_cache(self, path):
        """加载缓存的映射关系"""
        with open(path, 'r', encoding='utf-8') as f:
            if yaml:
                self.cache = yaml.safe_load(f) or {}
            else:
                self.cache = json.load(f)

        # 合并缓存的映射
        cached_mapping = self.cache.get('field_mapping', {})
        self.mapping.update(cached_mapping)

    def analyze_template(self, template_path=None):
        """
        分析模板结构

        Returns:
            dict: 模板分析结果
        """
        path = Path(template_path) if template_path else Path(self.template_path)
        if not path.exists():
            raise FileNotFoundError(f"模板文件不存在: {path}")

        wb = openpyxl.load_workbook(path)
        ws = wb.active

        # 读取表头（第一行）
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append({
                    'column': cell.column,
                    'letter': openpyxl.utils.get_column_letter(cell.column),
                    'value': str(cell.value).strip(),
                })

        return {
            'path': str(path),
            'sheet': ws.title,
            'headers': headers,
            'total_rows': ws.max_row,
            'total_cols': ws.max_column,
        }

    def auto_match_fields(self, template_headers):
        """
        自动匹配发票字段与模板字段

        Args:
            template_headers: 模板表头列表

        Returns:
            dict: 自动匹配结果
        """
        # 常见字段名映射
        common_mappings = {
            'invoice_date': ['开票日期', '日期', '发票日期', '发生日期'],
            'seller_name': ['销方名称', '销售方', '销方', '供应商', '收款方'],
            'amount': ['金额（不含税）', '金额', '不含税金额', '金额(不含税)'],
            'tax_amount': ['税额', '税金', '税款'],
            'total': ['价税合计', '合计', '价税合计(大写)', '合计金额'],
            'remark': ['费用说明', '报销事由', '备注', '说明'],
            'invoice_code': ['发票代码', '发票代号'],
            'invoice_number': ['发票号码', '发票号', '号码'],
            'expense_type': ['费用类型', '费用类别', '类别'],
        }

        matches = []
        for header in template_headers:
            header_value = header['value']
            for invoice_field, template_names in common_mappings.items():
                for template_name in template_names:
                    if self._fuzzy_match(header_value, template_name):
                        matches.append({
                            'invoice_field': invoice_field,
                            'template_field': header_value,
                            'column': header['letter'],
                            'confidence': 0.8 if header_value == template_name else 0.6,
                        })

        return matches

    def _fuzzy_match(self, s1, s2):
        """简单模糊匹配"""
        s1 = s1.strip().lower()
        s2 = s2.strip().lower()

        # 精确匹配
        if s1 == s2:
            return True

        # 包含匹配
        if s2 in s1 or s1 in s2:
            return True

        # 关键词匹配
        keywords = ['日期', '金额', '合计', '税额', '名称', '说明', '代码', '号码']
        for kw in keywords:
            if kw in s1 and kw in s2:
                return True

        return False

    def learn_from_correction(self, original_mapping, corrected_mapping, template_path=None):
        """
        从用户修正中学习

        Args:
            original_mapping: 原始映射
            corrected_mapping: 用户修正后的映射
            template_path: 模板路径（用于缓存）
        """
        # 更新缓存
        cache = self.cache.get('field_mapping', {})
        cache.update(corrected_mapping)

        # 保存缓存
        if self.config_path:
            cache_path = Path(self.config_path).parent / 'template_cache.yaml'
            cache_data = {
                'field_mapping': cache,
                'template_path': str(template_path or self.template_path),
                'learned_at': datetime.now().isoformat(),
            }
            with open(cache_path, 'w', encoding='utf-8') as f:
                if yaml:
                    yaml.dump(cache_data, f, allow_unicode=True, default_flow_style=False)
                else:
                    json.dump(cache_data, f, ensure_ascii=False, indent=2)

        return {
            'status': 'learned',
            'message': f'已学习 {len(corrected_mapping)} 条映射关系',
            'cache_path': str(cache_path) if self.config_path else None,
        }

    def fill_template(self, expense_data, template_path=None, output_path=None):
        """
        填充报销单模板

        Args:
            expense_data: 报销单数据dict
            template_path: 模板路径（可选，默认使用初始化时的模板）
            output_path: 输出路径（可选）

        Returns:
            dict: 填充结果
        """
        path = Path(template_path) if template_path else Path(self.template_path)
        if not path.exists():
            raise FileNotFoundError(f"模板文件不存在: {path}")

        wb = openpyxl.load_workbook(path)
        ws = wb.active

        # 构建字段位置映射（列号）
        header_map = {}
        for cell in ws[1]:
            if cell.value:
                header_map[str(cell.value).strip()] = cell.column

        # 填充数据
        filled = []
        for field_name, cell_col in header_map.items():
            # 查找发票字段映射
            invoice_field = None
            for inv_field, tmpl_field in self.mapping.items():
                if tmpl_field == field_name:
                    invoice_field = inv_field
                    break

            if not invoice_field:
                # 尝试直接匹配
                if field_name in expense_data:
                    invoice_field = field_name
                    # 更新映射
                    self.mapping[invoice_field] = field_name

            if invoice_field and invoice_field in expense_data:
                # 读取映射的字段名
                mapped_field = self.mapping.get(invoice_field, invoice_field)
                value = expense_data.get(mapped_field, expense_data.get(invoice_field))
                if value:
                    # 找到下一空行填充
                    next_row = ws.max_row + 1
                    ws.cell(row=next_row, column=cell_col, value=value)
                    filled.append({
                        'field': field_name,
                        'value': value,
                        'row': next_row,
                        'column': openpyxl.utils.get_column_letter(cell_col),
                    })

        # 保存
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = Path(path).parent / f"expense_report_{timestamp}.xlsx"

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        return {
            'success': True,
            'output_path': str(output_path),
            'filled_fields': len(filled),
            'filled_details': filled,
        }

    def merge_multiple_receipts(self, receipts, template_path=None, output_path=None):
        """
        合并多张发票到一张报销单

        Args:
            receipts: 发票数据列表
            template_path: 模板路径
            output_path: 输出路径

        Returns:
            dict: 合并结果
        """
        path = Path(template_path) if template_path else Path(self.template_path)
        if not path.exists():
            raise FileNotFoundError(f"模板文件不存在: {path}")

        wb = openpyxl.load_workbook(path)
        ws = wb.active

        # 构建字段位置映射
        header_map = {}
        for cell in ws[1]:
            if cell.value:
                header_map[str(cell.value).strip()] = cell.column

        # 逐行填充
        current_row = ws.max_row + 1
        total_amount = 0
        all_filled = []

        for receipt in receipts:
            # 计算合计
            total = receipt.get('total', 0)
            total_amount += total

            for field_name, cell_col in header_map.items():
                for inv_field, tmpl_field in self.mapping.items():
                    if tmpl_field == field_name and inv_field in receipt:
                        ws.cell(row=current_row, column=cell_col, value=receipt[inv_field])
                        all_filled.append({
                            'field': field_name,
                            'row': current_row,
                            'value': str(receipt[inv_field])[:50],
                        })
            current_row += 1

        # 保存
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = Path(path).parent / f"expense_merged_{timestamp}.xlsx"
        else:
            output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        return {
            'success': True,
            'output_path': str(output_path),
            'receipt_count': len(receipts),
            'total_amount': round(total_amount, 2),
            'filled_fields': len(all_filled),
        }


def main():
    """命令行入口"""
    parser = argparse.ArgumentParser(description='报销单模板引擎')
    subparsers = parser.add_subparsers(dest='command', help='子命令')

    # analyze子命令
    p_analyze = subparsers.add_parser('analyze', help='分析模板结构')
    p_analyze.add_argument('--template', required=True, help='模板文件路径')

    # fill子命令
    p_fill = subparsers.add_parser('fill', help='填充报销单')
    p_fill.add_argument('--config', required=True, help='配置文件路径')
    p_fill.add_argument('--receipt', required=True, help='发票数据JSON文件路径')
    p_fill.add_argument('--template', required=True, help='模板文件路径')
    p_fill.add_argument('--output', help='输出文件路径（可选）')

    # merge子命令
    p_merge = subparsers.add_parser('merge', help='合并多张发票')
    p_merge.add_argument('--config', required=True, help='配置文件路径')
    p_merge.add_argument('--receipts', required=True, help='发票数据JSON文件列表（逗号分隔）')
    p_merge.add_argument('--template', required=True, help='模板文件路径')
    p_merge.add_argument('--output', help='输出文件路径（可选）')

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        sys.exit(1)

    if args.command == 'analyze':
        matcher = TemplateMatcher(template_path=args.template)
        result = matcher.analyze_template()
        print(json.dumps(result, ensure_ascii=False, indent=2))

    elif args.command == 'fill':
        matcher = TemplateMatcher(template_path=args.template, config_path=args.config)
        with open(args.receipt, 'r', encoding='utf-8') as f:
            expense_data = json.load(f)
        result = matcher.fill_template(expense_data, output_path=args.output)
        print(json.dumps(result, ensure_ascii=False, indent=2))

    elif args.command == 'merge':
        matcher = TemplateMatcher(template_path=args.template, config_path=args.config)
        receipt_paths = args.receipts.split(',')
        receipts = []
        for path in receipt_paths:
            path = path.strip()
            if Path(path).exists():
                with open(path, 'r', encoding='utf-8') as f:
                    receipts.append(json.load(f))
        result = matcher.merge_multiple_receipts(receipts, output_path=args.output)
        print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
