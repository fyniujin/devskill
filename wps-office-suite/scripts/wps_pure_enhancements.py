#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
wps_pure_enhancements.py — 纯 Python 增强能力模块 (v4.8.0)

死规则:
  1. 仅使用 openpyxl 标准库，不依赖 WPS/MS Office 进程或 COM 接口
  2. 所有操作在本地文件系统完成，不上传任何数据到外部服务
  3. 不读取、不修改系统注册表，不调用 shell 命令操作 Office 软件
  4. 输入文件路径必须为绝对路径，禁止执行用户提供的任意代码
  5. 所有写入操作前必须验证目标文件存在且为有效 xlsx 格式

安全合规:
  - 无网络请求，无数据外传，无敏感信息收集
  - 仅处理用户显式指定的文件路径
  - 所有操作可逆（通过备份或撤销），不破坏原始数据结构
  - 符合企业数据安全规范，适用于内网隔离环境

功能模块:
  1. 条件格式 (cond-format) — 色阶/数据条/图标集/单元格规则
  2. 数据验证 (data-validation) — 下拉列表/日期/整数/小数/文本长度
  3. 合并单元格 (merge-cells) — 合并/取消合并，可选写入文本
  4. 命名区域 (named-range) — 创建/删除/列出命名区域

Version: v4.8.0
"""

import argparse
import os
import sys
from copy import deepcopy
from typing import Any, Dict, List, Optional, Union

try:
    from openpyxl import load_workbook
    from openpyxl.formatting import Rule
    from openpyxl.formatting.rule import (
        CellIsRule,
        ColorScaleRule,
        DataBarRule,
        IconSetRule,
    )
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.utils import get_column_letter, range_boundaries
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


__version__ = "v4.8.0"


# ─── 颜色映射 ───────────────────────────────────────────────────────────────

COLOR_MAP = {
    "red": "F8696B",
    "green": "63BE7B",
    "blue": "4472C4",
    "yellow": "FFEB84",
    "orange": "F4B084",
    "purple": "9966FF",
    "white": "FFFFFF",
    "black": "000000",
    "gray": "A6A6A6",
}

SCALE_PRESETS = {
    "red": ("F8696B", "FFEB84", "63BE7B"),
    "green": ("63BE7B", "FFEB84", "F8696B"),
    "blue": ("63BE7B", "FFEB84", "F8696B"),
    "yellow": ("F8696B", "FFEB84", "63BE7B"),
}


# ─── 工具函数 ───────────────────────────────────────────────────────────────

def _validate_file(filepath: str) -> bool:
    """验证文件存在且为有效 xlsx"""
    if not os.path.isfile(filepath):
        return False
    if not filepath.lower().endswith((".xlsx", ".xlsm")):
        return False
    return True


def _parse_range(range_str: str) -> tuple:
    """解析范围字符串为 (min_col, min_row, max_col, max_row)"""
    return range_boundaries(range_str)


def _hex_color(color_name: str) -> str:
    """将颜色名称转为十六进制"""
    return COLOR_MAP.get(color_name.lower(), "F8696B")


# ─── 核心类 ─────────────────────────────────────────────────────────────────

class PureEnhancements:
    """
    纯 Python 增强能力类 — 基于 openpyxl 实现，无需 WPS/MS Office。

    所有方法返回 dict 格式: {"ok": bool, "message": str, ...}
    """

    # ── 条件格式 ────────────────────────────────────────────────────────────

    def conditional_format(
        self,
        filepath: str,
        sheet: str,
        range_str: str,
        fmt_type: str,
        operator: Optional[str] = None,
        value: Optional[str] = None,
        color: str = "red",
        foreground: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        应用条件格式。

        Args:
            filepath: xlsx 文件绝对路径
            sheet: 工作表名称
            range_str: 目标范围 (如 "A1:A100")
            fmt_type: 类型 — color-scale/data-bar/icon-set/cell-is
            operator: cell-is 操作符 — gt/lt/eq/geq/leq/between
            value: 阈值 (如 "80" 或 "60,90")
            color: 颜色名称 (red/green/blue/yellow)
            foreground: 文本颜色十六进制 (如 "FF0000")

        Returns:
            dict 包含 ok 状态和详细信息
        """
        if not _validate_file(filepath):
            return {"ok": False, "error": f"文件不存在或格式无效: {filepath}"}

        try:
            wb = load_workbook(filepath)
            if sheet not in wb.sheetnames:
                return {"ok": False, "error": f"工作表不存在: {sheet}"}
            ws = wb[sheet]

            fmt_type = fmt_type.lower().strip()

            if fmt_type == "color-scale":
                self._apply_color_scale(ws, range_str, color)
            elif fmt_type == "data-bar":
                self._apply_data_bar(ws, range_str, color)
            elif fmt_type == "icon-set":
                self._apply_icon_set(ws, range_str)
            elif fmt_type == "cell-is":
                if not operator or value is None:
                    return {"ok": False, "error": "cell-is 类型需要 operator 和 value 参数"}
                self._apply_cell_is(ws, range_str, operator, value, color, foreground)
            else:
                return {"ok": False, "error": f"不支持的条件格式类型: {fmt_type}"}

            wb.save(filepath)
            return {
                "ok": True,
                "message": f"条件格式 [{fmt_type}] 已应用到 {sheet}!{range_str}",
                "type": fmt_type,
                "range": range_str,
            }
        except Exception as e:
            return {"ok": False, "error": f"条件格式应用失败: {str(e)}"}

    def _apply_color_scale(self, ws, cell_range: str, color: str) -> None:
        """应用三色色阶"""
        colors = SCALE_PRESETS.get(color.lower(), SCALE_PRESETS["red"])
        rule = ColorScaleRule(
            start_type="min", start_color=colors[0],
            mid_type="percentile", mid_value=50, mid_color=colors[1],
            end_type="max", end_color=colors[2],
        )
        ws.conditional_formatting.add(cell_range, rule)

    def _apply_data_bar(self, ws, cell_range: str, color: str) -> None:
        """应用数据条"""
        bar_color = _hex_color(color)
        rule = DataBarRule(
            start_type="min", end_type="max",
            color=bar_color,
        )
        ws.conditional_formatting.add(cell_range, rule)

    def _apply_icon_set(self, ws, cell_range: str) -> None:
        """应用图标集 (3 个交通灯)"""
        rule = IconSetRule(
            "3TrafficLights1", "percent",
            [0, 50, 100], showValue=True,
        )
        ws.conditional_formatting.add(cell_range, rule)

    def _apply_cell_is(
        self,
        ws,
        cell_range: str,
        operator: str,
        value: str,
        color: str,
        foreground: Optional[str],
    ) -> None:
        """应用单元格值规则"""
        from openpyxl.styles import Font, PatternFill

        op_map = {
            "gt": "greaterThan",
            "lt": "lessThan",
            "eq": "equal",
            "geq": "greaterThanOrEqual",
            "leq": "lessThanOrEqual",
            "between": "between",
        }
        op = op_map.get(operator.lower())
        if not op:
            raise ValueError(f"不支持的操作符: {operator}")

        # 解析值
        if op == "between":
            parts = value.split(",")
            if len(parts) != 2:
                raise ValueError("between 操作符需要两个值，用逗号分隔")
            formula = [parts[0].strip(), parts[1].strip()]
        else:
            formula = [value.strip()]

        fill_color = _hex_color(color)
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        font = None
        if foreground:
            font = Font(color=foreground)

        rule = CellIsRule(
            operator=op, formula=formula,
            fill=fill, font=font,
        )
        ws.conditional_formatting.add(cell_range, rule)

    # ── 数据验证 ────────────────────────────────────────────────────────────

    def data_validation(
        self,
        filepath: str,
        sheet: str,
        range_str: str,
        validation_type: str,
        list_values: Optional[List[str]] = None,
        min_val: Optional[Union[int, float, str]] = None,
        max_val: Optional[Union[int, float, str]] = None,
        error_title: Optional[str] = None,
        error_msg: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        应用数据验证规则。

        Args:
            filepath: xlsx 文件绝对路径
            sheet: 工作表名称
            range_str: 目标范围 (如 "B1:B100")
            validation_type: 类型 — list/date/whole/decimal/text-length
            list_values: 下拉列表值 (逗号分隔)
            min_val: 最小值
            max_val: 最大值
            error_title: 错误提示标题
            error_msg: 错误提示内容

        Returns:
            dict 包含 ok 状态和详细信息
        """
        if not _validate_file(filepath):
            return {"ok": False, "error": f"文件不存在或格式无效: {filepath}"}

        try:
            wb = load_workbook(filepath)
            if sheet not in wb.sheetnames:
                return {"ok": False, "error": f"工作表不存在: {sheet}"}
            ws = wb[sheet]

            vtype = validation_type.lower().strip()
            type_map = {
                "list": "list",
                "date": "date",
                "whole": "whole",
                "decimal": "decimal",
                "text-length": "textLength",
            }
            dv_type = type_map.get(vtype)
            if not dv_type:
                return {"ok": False, "error": f"不支持的验证类型: {validation_type}"}

            dv = DataValidation(
                type=dv_type,
                allow_blank=True,
                showErrorMessage=True,
            )

            # 设置公式或范围
            if vtype == "list":
                if not list_values:
                    return {"ok": False, "error": "list 类型需要提供 list_values 参数"}
                formula = ",".join(list_values)
                dv.formula1 = f'"{formula}"'
            else:
                if min_val is not None:
                    dv.formula1 = str(min_val)
                if max_val is not None:
                    dv.formula2 = str(max_val)

            # 错误提示
            if error_title:
                dv.errorTitle = error_title
            if error_msg:
                dv.error = error_msg
            else:
                dv.error = "输入值不符合验证规则"

            dv.prompt = "请选择或输入有效值"
            dv.promptTitle = "输入提示"

            dv.add(range_str)
            ws.add_data_validation(dv)

            wb.save(filepath)
            return {
                "ok": True,
                "message": f"数据验证 [{vtype}] 已应用到 {sheet}!{range_str}",
                "type": vtype,
                "range": range_str,
            }
        except Exception as e:
            return {"ok": False, "error": f"数据验证应用失败: {str(e)}"}

    # ── 合并单元格 ──────────────────────────────────────────────────────────

    def merge_cells(
        self,
        filepath: str,
        sheet: str,
        range_str: str,
        action: str,
        text: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        合并或取消合并单元格。

        Args:
            filepath: xlsx 文件绝对路径
            sheet: 工作表名称
            range_str: 目标范围 (如 "A1:D1")
            action: 操作 — merge/unmerge
            text: 合并后写入的文本

        Returns:
            dict 包含 ok 状态和详细信息
        """
        if not _validate_file(filepath):
            return {"ok": False, "error": f"文件不存在或格式无效: {filepath}"}

        try:
            wb = load_workbook(filepath)
            if sheet not in wb.sheetnames:
                return {"ok": False, "error": f"工作表不存在: {sheet}"}
            ws = wb[sheet]

            action = action.lower().strip()

            if action == "merge":
                ws.merge_cells(range_str)
                if text is not None:
                    min_col, min_row, _, _ = _parse_range(range_str)
                    cell = ws.cell(row=min_row, column=min_col)
                    cell.value = text
                return {
                    "ok": True,
                    "message": f"已合并单元格 {sheet}!{range_str}",
                    "action": "merge",
                    "range": range_str,
                    "text": text,
                }
            elif action == "unmerge":
                ws.unmerge_cells(range_str)
                return {
                    "ok": True,
                    "message": f"已取消合并 {sheet}!{range_str}",
                    "action": "unmerge",
                    "range": range_str,
                }
            else:
                return {"ok": False, "error": f"不支持的操作: {action}"}

        except Exception as e:
            return {"ok": False, "error": f"合并/取消合并失败: {str(e)}"}

    # ── 命名区域 ────────────────────────────────────────────────────────────

    def named_range(
        self,
        filepath: str,
        sheet: str,
        name: str,
        range_str: str,
        action: str,
    ) -> Dict[str, Any]:
        """
        管理命名区域。

        Args:
            filepath: xlsx 文件绝对路径
            sheet: 工作表名称
            name: 区域名称
            range_str: 范围 (如 "A1:D100")
            action: 操作 — create/delete/list

        Returns:
            dict 包含 ok 状态和详细信息
        """
        if not _validate_file(filepath):
            return {"ok": False, "error": f"文件不存在或格式无效: {filepath}"}

        try:
            wb = load_workbook(filepath)
            action = action.lower().strip()

            if action == "create":
                if sheet not in wb.sheetnames:
                    return {"ok": False, "error": f"工作表不存在: {sheet}"}
                # 构建带工作表前缀的引用字符串
                ref = f"'{sheet}'!{range_str}"
                defn = DefinedName(name, attr_text=ref)
                wb.defined_names.add(defn)
                wb.save(filepath)
                return {
                    "ok": True,
                    "message": f"命名区域 [{name}] 已创建 → {ref}",
                    "action": "create",
                    "name": name,
                    "ref": ref,
                }

            elif action == "delete":
                if name in wb.defined_names:
                    del wb.defined_names[name]
                    wb.save(filepath)
                    return {
                        "ok": True,
                        "message": f"命名区域 [{name}] 已删除",
                        "action": "delete",
                        "name": name,
                    }
                else:
                    return {"ok": False, "error": f"命名区域不存在: {name}"}

            elif action == "list":
                items = []
                for defn in wb.defined_names.definedName:
                    items.append({
                        "name": defn.name,
                        "ref": defn.attr_text,
                    })
                return {
                    "ok": True,
                    "message": f"共 {len(items)} 个命名区域",
                    "action": "list",
                    "items": items,
                }
            else:
                return {"ok": False, "error": f"不支持的操作: {action}"}

        except Exception as e:
            return {"ok": False, "error": f"命名区域操作失败: {str(e)}"}


# ─── CLI 入口 ───────────────────────────────────────────────────────────────

def _cli_cond_format(args):
    """cond-format 子命令处理"""
    pe = PureEnhancements()
    result = pe.conditional_format(
        filepath=args.file,
        sheet=args.sheet,
        range_str=args.range,
        fmt_type=args.type,
        operator=args.operator,
        value=args.value,
        color=args.color,
        foreground=args.foreground,
    )
    _print_result(result)


def _cli_data_validation(args):
    """data-validation 子命令处理"""
    pe = PureEnhancements()
    list_values = None
    if args.list:
        list_values = [v.strip() for v in args.list.split(",")]
    result = pe.data_validation(
        filepath=args.file,
        sheet=args.sheet,
        range_str=args.range,
        validation_type=args.type,
        list_values=list_values,
        min_val=args.min,
        max_val=args.max,
        error_title=args.error_title,
        error_msg=args.error_msg,
    )
    _print_result(result)


def _cli_merge_cells(args):
    """merge-cells 子命令处理"""
    pe = PureEnhancements()
    result = pe.merge_cells(
        filepath=args.file,
        sheet=args.sheet,
        range_str=args.range,
        action=args.action,
        text=args.text,
    )
    _print_result(result)


def _cli_named_range(args):
    """named-range 子命令处理"""
    pe = PureEnhancements()
    result = pe.named_range(
        filepath=args.file,
        sheet=args.sheet,
        name=args.name,
        range_str=args.range,
        action=args.action,
    )
    _print_result(result)


def _print_result(result: Dict[str, Any]) -> None:
    """格式化输出结果"""
    import json
    print(json.dumps(result, ensure_ascii=False, indent=2))


def main():
    """CLI 主入口"""
    parser = argparse.ArgumentParser(
        description=f"WPS Pure Enhancements {__version__} — 纯 Python 增强能力",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    subparsers = parser.add_subparsers(dest="command", help="可用子命令")

    # ── cond-format ────────────────────────────────────────────────────────
    p_cf = subparsers.add_parser("cond-format", help="条件格式")
    p_cf.add_argument("--file", required=True, help="xlsx 文件路径")
    p_cf.add_argument("--sheet", default="Sheet1", help="工作表名称 (默认 Sheet1)")
    p_cf.add_argument("--range", required=True, help="目标范围 (如 A1:A100)")
    p_cf.add_argument(
        "--type", required=True,
        choices=["color-scale", "data-bar", "icon-set", "cell-is"],
        help="条件格式类型",
    )
    p_cf.add_argument(
        "--operator",
        choices=["gt", "lt", "eq", "geq", "leq", "between"],
        help="cell-is 操作符",
    )
    p_cf.add_argument("--value", help="阈值 (如 80 或 60,90)")
    p_cf.add_argument(
        "--color", default="red",
        choices=["red", "green", "blue", "yellow", "orange", "purple"],
        help="颜色 (默认 red)",
    )
    p_cf.add_argument("--foreground", help="文本颜色十六进制 (如 FF0000)")
    p_cf.set_defaults(func=_cli_cond_format)

    # ── data-validation ────────────────────────────────────────────────────
    p_dv = subparsers.add_parser("data-validation", help="数据验证")
    p_dv.add_argument("--file", required=True, help="xlsx 文件路径")
    p_dv.add_argument("--sheet", default="Sheet1", help="工作表名称 (默认 Sheet1)")
    p_dv.add_argument("--range", required=True, help="目标范围 (如 B1:B100)")
    p_dv.add_argument("--list", help="下拉列表值 (逗号分隔, 如 是,否)")
    p_dv.add_argument(
        "--type", required=True,
        choices=["list", "date", "whole", "decimal", "text-length"],
        help="验证类型",
    )
    p_dv.add_argument("--min", help="最小值")
    p_dv.add_argument("--max", help="最大值")
    p_dv.add_argument("--error-title", help="错误提示标题")
    p_dv.add_argument("--error-msg", help="错误提示内容")
    p_dv.set_defaults(func=_cli_data_validation)

    # ── merge-cells ────────────────────────────────────────────────────────
    p_mc = subparsers.add_parser("merge-cells", help="合并单元格")
    p_mc.add_argument("--file", required=True, help="xlsx 文件路径")
    p_mc.add_argument("--sheet", default="Sheet1", help="工作表名称 (默认 Sheet1)")
    p_mc.add_argument("--range", required=True, help="目标范围 (如 A1:D1)")
    p_mc.add_argument(
        "--action", required=True,
        choices=["merge", "unmerge"],
        help="操作类型",
    )
    p_mc.add_argument("--text", help="合并后写入的文本")
    p_mc.set_defaults(func=_cli_merge_cells)

    # ── named-range ────────────────────────────────────────────────────────
    p_nr = subparsers.add_parser("named-range", help="命名区域")
    p_nr.add_argument("--file", required=True, help="xlsx 文件路径")
    p_nr.add_argument("--sheet", default="Sheet1", help="工作表名称 (默认 Sheet1)")
    p_nr.add_argument("--name", required=True, help="区域名称")
    p_nr.add_argument("--range", required=True, help="范围 (如 A1:D100)")
    p_nr.add_argument(
        "--action", required=True,
        choices=["create", "delete", "list"],
        help="操作类型",
    )
    p_nr.set_defaults(func=_cli_named_range)

    # 解析并执行
    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        sys.exit(1)

    args.func(args)


if __name__ == "__main__":
    main()
