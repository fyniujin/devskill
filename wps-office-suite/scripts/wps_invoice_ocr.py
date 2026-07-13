"""
Excel 发票 OCR 入账器 v4.0
输入 PDF/图片发票 → OCR 提取字段 → 自动填充到 Excel 表格

核心能力（本地 OCR，不依赖 API）：
  • PDF 解析（pdfplumber）+ 图片 OCR（pytesseract）
  • 正则表达式提取：发票号、日期、金额、税号、收款方
  • 自动分类入账（按日期/供应商/金额）
  • 纯 Python 实现，支持批量处理

依赖安装：
  pip install pytesseract pdfplumber openpyxl
  # 安装 Tesseract OCR 引擎（Windows 安装包）：
  # https://github.com/UB-Mannheim/tesseract/wiki
"""
import re
import json
from pathlib import Path
from typing import List, Dict, Optional
from datetime import datetime


# ==================== 发票字段提取 ====================

# 通用正则模式（适配中国增值税发票）
INVOICE_PATTERNS = {
    "invoice_number": [
        r"发票代码[：:\s]*([0-9]{10,12})",
        r"发票号码[：:\s]*([0-9]{8,12})",
        r"No[:\s]*([0-9]{8,12})",
        r"([0-9]{8,12})\s*号",
    ],
    "date": [
        r"开票日期[：:\s]*(\d{4}年\d{1,2}月\d{1,2})",
        r"日期[：:\s]*(\d{4}[-/年]\d{1,2}[-/月]\d{1,2})",
        r"(\d{4}[-/]\d{1,2}[-/]\d{1,2})",
    ],
    "total_amount": [
        r"价税合计[（(]小写[）)][：:\s]*[¥￥]?\s*([0-9,.]+)",
        r"小写[：:\s]*[¥￥]?\s*([0-9,.]+)",
        r"合计[：:\s]*[¥￥]?\s*([0-9,.]+)",
        r"金额[：:\s]*[¥￥]?\s*([0-9,.]+)",
        r"[¥￥]\s*([0-9,.]+)",
    ],
    "tax_number": [
        r"纳税人识别号[：:\s]*([0-9A-Z]{18,20})",
        r"统一社会信用代码[：:\s]*([0-9A-Z]{18})",
        r"税号[：:\s]*([0-9A-Z]{18,20})",
    ],
    "seller_name": [
        r"收款方[：:\s]*([^\n]+?公司)",
        r"销售方[：:\s]*([^\n]+?公司)",
        r"名称[：:\s]*([^\n]+?公司)",
    ],
    "buyer_name": [
        r"购买方[：:\s]*([^\n]+?公司)",
        r"付款方[：:\s]*([^\n]+?公司)",
    ],
    "tax_amount": [
        r"税额[：:\s]*[¥￥]?\s*([0-9,.]+)",
        r"税[：:\s]*([0-9,.]+)[%％]",
    ],
}

# 常见发票类型
INVOICE_TYPES = {
    "增值税专用发票": "专票",
    "增值税普通发票": "普票",
    "电子发票": "电子",
    "机动车": "机动车",
    "通行费": "通行费",
    "出租车": "出租车",
}


def extract_text_from_pdf(filepath: str) -> str:
    """从 PDF 提取文本（pdfplumber）"""
    try:
        import pdfplumber
        text_parts = []
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
        return "\n".join(text_parts)
    except ImportError:
        return ""


def extract_text_from_image(image_path: str) -> str:
    """从图片提取文本（pytesseract）"""
    try:
        import pytesseract
        from PIL import Image
        img = Image.open(image_path)
        text = pytesseract.image_to_string(img, lang="chi_sim+eng")
        return text
    except ImportError:
        return ""


def parse_invoice_text(raw_text: str) -> Dict:
    """从原始文本中提取发票字段"""
    result = {}

    for field, patterns in INVOICE_PATTERNS.items():
        for pattern in patterns:
            match = re.search(pattern, raw_text)
            if match:
                value = match.group(1).strip() if match.lastindex else match.group(0).strip()
                # 清理金额中的逗号
                if "amount" in field:
                    value = value.replace(",", "")
                result[field] = value
                break

    # 确定发票类型
    for type_name, short_name in INVOICE_TYPES.items():
        if type_name in raw_text:
            result["invoice_type"] = type_name
            break

    return result


def process_invoice(input_path: str) -> Dict:
    """处理单个发票文件，返回提取的字段"""
    try:
        input_p = Path(input_path)
        if not input_p.exists():
            return {"success": False, "error": f"文件不存在: {input_path}"}

        ext = input_p.suffix.lower()
        raw_text = ""

        if ext == ".pdf":
            raw_text = extract_text_from_pdf(input_path)
        elif ext in (".jpg", ".jpeg", ".png", ".bmp", ".tiff"):
            raw_text = extract_text_from_image(input_path)
        else:
            return {"success": False, "error": f"不支持的文件格式: {ext}"}

        if not raw_text:
            return {"success": False, "error": "OCR 未能提取到文本，请检查文件质量或安装 OCR 引擎"}

        fields = parse_invoice_text(raw_text)
        fields["source_file"] = input_p.name
        fields["processed_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        fields["text_preview"] = raw_text[:200]

        return {"success": True, "fields": fields}

    except Exception as e:
        return {"success": False, "error": str(e)}


# ==================== 入账到 Excel ====================

def generate_accounting_template(output_path: str) -> Dict:
    """生成入账模板 Excel 文件"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "发票入账"

        # 表头
        headers = [
            "序号", "发票类型", "发票代码", "发票号码",
            "开票日期", "金额", "税额", "价税合计",
            "购买方", "销售方", "税率", "备注",
            "入账日期", "附件文件名"
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1A73E8")

        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # 列宽
        col_widths = [6, 12, 12, 14, 12, 14, 14, 14, 20, 20, 8, 14, 12, 25]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

        wb.save(output_path)
        return {"success": True, "path": output_path}
    except ImportError:
        return {"success": False, "error": "需要安装 openpyxl"}


def invoice_to_excel(invoice_result: dict, output_path: str = None) -> Dict:
    """将提取的发票信息写入入账 Excel"""
    try:
        import openpyxl

        fields = invoice_result.get("fields", {})
        if not fields:
            return {"success": False, "error": "没有提取到发票字段"}

        if not output_path:
            output_path = str(Path(invoice_result.get("fields", {}).get("source_file", "未命名发票")).parent / "发票入账台账.xlsx")

        # 判断是新建还是追加
        if Path(output_path).exists():
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
        else:
            template_result = generate_accounting_template(output_path)
            if not template_result["success"]:
                return template_result
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active

        # 找下一行
        next_row = ws.max_row + 1
        if next_row == 2:
            seq = 1
        else:
            seq = next_row - 1

        # 写入数据（按入账表顺序）
        data = {
            1: seq,
            2: fields.get("invoice_type", "未识别"),
            3: fields.get("invoice_code", fields.get("invoice_number", "")),
            4: fields.get("invoice_number", ""),
            5: fields.get("date", ""),
            6: fields.get("total_amount", fields.get("total_amount", "")),
            7: fields.get("tax_amount", ""),
            8: fields.get("total_amount", ""),
            9: fields.get("buyer_name", ""),
            10: fields.get("seller_name", ""),
            11: "13%",  # 默认税率，可后续修改
            12: "",
            13: datetime.now().strftime("%Y-%m-%d"),
            14: fields.get("source_file", ""),
        }

        for c, val in data.items():
            ws.cell(row=next_row, column=c, value=val)

        # 添加边框和格式
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for col in range(1, 15):
            ws.cell(row=next_row, column=col).border = thin_border

        wb.save(output_path)

        return {
            "success": True,
            "path": output_path,
            "next_row": next_row,
            "fields_summary": f"发票号: {fields.get('invoice_number', '未识别')}，金额: ¥{fields.get('total_amount', '0')}，日期: {fields.get('date', '未知')}",
        }

    except ImportError:
        return {"success": False, "error": "需要安装 openpyxl"}
    except Exception as e:
        return {"success": False, "error": str(e)}


# ==================== 批量入账 ====================

def batch_invoices_to_accounting(input_paths: List[str], output_path: str = None) -> Dict:
    """批量处理发票并入账"""
    results = []
    success_count = 0

    for input_path in input_paths:
        result = process_invoice(input_path)
        if result["success"]:
            # 入账到同一个 Excel
            entry_result = invoice_to_excel(result, output_path)
            result["entry"] = entry_result
            if entry_result["success"]:
                success_count += 1
            output_path = output_path or entry_result.get("path")
        results.append(result)

    return {
        "success": True,
        "total": len(input_paths),
        "success_count": success_count,
        "fail_count": len(input_paths) - success_count,
        "output": output_path,
        "details": results,
    }


# ==================== 通用入口 ====================

def invoice_to_accounting(input_path: str, output_path: str = None) -> Dict:
    """
    单张发票入账

    Args:
        input_path: 发票文件路径（PDF/图片）
        output_path: 输出 Excel 路径（可选，默认同目录）
    """
    result = process_invoice(input_path)
    if result["success"]:
        entry_result = invoice_to_excel(result, output_path)
        result["entry"] = entry_result
        result["accounting_file"] = entry_result.get("path")
    return result


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="发票 OCR 入账")
    sub = parser.add_subparsers(dest="command", required=True)

    p_single = sub.add_parser("single", help="单张发票入账")
    p_single.add_argument("--input", required=True, help="发票文件路径（PDF/图片）")
    p_single.add_argument("--output", default="", help="输出 Excel 路径")

    p_batch = sub.add_parser("batch", help="批量发票入账")
    p_batch.add_argument("--inputs", required=True, help="多个文件路径，逗号分隔")
    p_batch.add_argument("--output", default="", help="输出 Excel 路径")

    p_ocr = sub.add_parser("ocr", help="仅 OCR 提取，不入账")
    p_ocr.add_argument("--input", required=True, help="发票文件路径")

    p_template = sub.add_parser("template", help="生成入账模板")
    p_template.add_argument("--output", required=True, help="模板文件输出路径")

    args = parser.parse_args()

    if args.command == "single":
        result = invoice_to_accounting(args.input, args.output)
    elif args.command == "batch":
        input_list = [p.strip() for p in args.inputs.split(",")]
        result = batch_invoices_to_accounting(input_list, args.output)
    elif args.command == "ocr":
        result = process_invoice(args.input)
    elif args.command == "template":
        result = generate_accounting_template(args.output)

    print(json.dumps(result, ensure_ascii=False, default=str))
