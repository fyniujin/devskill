"""
公式解释器 v4.7.0
功能：Excel 公式→自然语言解释（纯本地实现，无外部 API）

v4.7.0 变更:
  - 🎯 Excel 公式解析为 AST
  - 🎯 AST→中文自然语言描述
  - 🎯 与 nl2formula 互逆（NL2Formula 的反向功能）
  - 🎯 批量解释 Sheet 中所有公式
  - 🎯 纯本地实现，不读取外部凭证或 API Key
"""

import os
import sys
import json
import re
from pathlib import Path
from typing import Dict, List, Optional, Any

sys.path.insert(0, str(Path(__file__).parent))

try:
    from wps_common import safe_path
except ImportError:
    def safe_path(p): return Path(p)

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class FormulaExplainer:
    """公式解释器（纯本地实现）"""
    
    # 函数中英文名称映射
    FUNC_NAMES = {
        "SUM": "求和",
        "AVERAGE": "平均值",
        "AVERAGEA": "平均值（含文本）",
        "COUNT": "计数",
        "COUNTA": "非空计数",
        "COUNTIF": "条件计数",
        "COUNTIFS": "多条件计数",
        "SUMIF": "条件求和",
        "SUMIFS": "多条件求和",
        "AVERAGEIF": "条件平均",
        "AVERAGEIFS": "多条件平均",
        "MAX": "最大值",
        "MIN": "最小值",
        "MEDIAN": "中位数",
        "MODE": "众数",
        "STDEV": "标准偏差",
        "VAR": "方差",
        "LARGE": "第K大值",
        "SMALL": "第K小值",
        "RANK": "排名",
        "QUARTILE": "四分位数",
        "PERCENTILE": "百分位数",
        "IF": "条件判断",
        "IFS": "多条件判断",
        "IFERROR": "错误时返回",
        "IFNA": "空值时返回",
        "AND": "逻辑与",
        "OR": "逻辑或",
        "NOT": "逻辑非",
        "XOR": "逻辑异或",
        "TRUE": "真",
        "FALSE": "假",
        "ISBLANK": "是否为空",
        "ISERROR": "是否为错误",
        "ISNA": "是否为#N/A",
        "ISTEXT": "是否为文本",
        "ISNUMBER": "是否为数字",
        "ISLOGICAL": "是否为逻辑值",
        "ISREF": "是否为引用",
        "VLOOKUP": "垂直查找",
        "HLOOKUP": "水平查找",
        "INDEX": "索引取值",
        "MATCH": "匹配位置",
        "OFFSET": "偏移引用",
        "INDIRECT": "间接引用",
        "ADDRESS": "地址文本",
        "ROW": "行号",
        "COLUMN": "列号",
        "ROWS": "行数",
        "COLUMNS": "列数",
        "CHOOSE": "选择值",
        "HYPERLINK": "超链接",
        "TRANSPOSE": "转置",
        "UPPER": "大写",
        "LOWER": "小写",
        "PROPER": "首字母大写",
        "LEN": "长度",
        "TRIM": "去除空格",
        "CLEAN": "清除不可打印字符",
        "CONCATENATE": "文本连接",
        "CONCAT": "文本连接",
        "TEXTJOIN": "文本连接（含分隔符）",
        "LEFT": "左取字符",
        "RIGHT": "右取字符",
        "MID": "中间取字符",
        "FIND": "查找位置（区分大小写）",
        "SEARCH": "查找位置（不区分大小写）",
        "REPLACE": "替换文本",
        "SUBSTITUTE": "替换文本（指定次数）",
        "TEXT": "数字格式化",
        "VALUE": "文本转数字",
        "ROUND": "四舍五入",
        "ROUNDUP": "向上舍入",
        "ROUNDDOWN": "向下舍入",
        "INT": "取整",
        "MOD": "取余",
        "POWER": "幂运算",
        "SQRT": "平方根",
        "ABS": "绝对值",
        "SIGN": "符号",
        "EXP": "指数",
        "LN": "自然对数",
        "LOG": "对数",
        "LOG10": "以10为底对数",
        "PI": "圆周率",
        "RAND": "随机数",
        "RANDBETWEEN": "区间随机数",
        "TODAY": "当前日期",
        "NOW": "当前时间",
        "DATE": "日期",
        "YEAR": "年份",
        "MONTH": "月份",
        "DAY": "日",
        "WEEKDAY": "星期几",
        "HOUR": "小时",
        "MINUTE": "分钟",
        "SECOND": "秒",
        "DATEDIF": "日期差",
        "EDATE": "月份偏移日期",
        "EOMONTH": "月末日期",
        "NETWORKDAYS": "工作日天数",
        "WORKDAY": "工作日偏移",
        "SUMPRODUCT": "数组乘积求和",
        "SUBTOTAL": "分类汇总",
        "AGGREGATE": "聚合",
        "CORREL": "相关系数",
        "COVAR": "协方差",
        "FORECAST": "预测",
        "GROWTH": "增长预测",
        "TREND": "趋势",
        "LINEST": "线性回归",
        "LOGEST": "对数回归",
        "SLOPE": "斜率",
        "INTERCEPT": "截距",
    }
    
    # 运算符映射
    OP_NAMES = {
        "+": "加上",
        "-": "减去",
        "*": "乘以",
        "/": "除以",
        "^": "的次幂",
        "&": "连接",
        "=": "等于",
        "<>": "不等于",
        ">": "大于",
        "<": "小于",
        ">=": "大于等于",
        "<=": "小于等于",
    }
    
    def __init__(self):
        pass
    
    def explain(self, formula: str) -> str:
        """解释 Excel 公式为自然语言"""
        if not formula:
            return "空公式"
        
        # 移除开头的 =
        formula = formula.lstrip()
        if formula.startswith("="):
            formula = formula[1:]
        
        # 移除首尾空格
        formula = formula.strip()
        
        if not formula:
            return "空公式"
        
        # 解析并生成解释
        try:
            return self._parse_and_explain(formula)
        except Exception:
            return f"无法解析的公式: {formula}"
    
    def _parse_and_explain(self, formula: str) -> str:
        """递归解析公式并生成解释"""
        formula = formula.strip()
        
        # 处理括号
        if formula.startswith("(") and formula.endswith(")"):
            inner = formula[1:-1]
            # 检查是否匹配
            depth = 0
            for ch in inner:
                if ch == "(":
                    depth += 1
                elif ch == ")":
                    depth -= 1
                if depth < 0:
                    break
            if depth == 0:
                return self._parse_and_explain(inner)
        
        # 处理字符串常量
        if formula.startswith('"') and formula.endswith('"'):
            return f"文本常量: {formula[1:-1]}"
        
        # 处理数字常量
        if re.match(r'^-?[\d.]+$', formula):
            return f"数值: {formula}"
        
        # 处理单元格引用
        if re.match(r'^[A-Z]+[0-9]+$', formula):
            return f"单元格 {formula} 的值"
        
        # 处理区域引用
        if re.match(r'^[A-Z]+[0-9]+:[A-Z]+[0-9]+$', formula):
            return f"区域 {formula}"
        
        # 处理带表名的引用
        if re.match(r'^[^!]+![A-Z]+[0-9]+$', formula):
            return f"表 {formula.replace('!', ' 的 ')}"
        if re.match(r'^[^!]+![A-Z]+[0-9]+:[A-Z]+[0-9]+$', formula):
            parts = formula.split('!')
            return f"表 {parts[0]} 的区域 {parts[1]}"
        
        # 尝试解析函数调用
        func_match = re.match(r'^([A-Z][A-Z0-9]*)\((.*)\)$', formula, re.DOTALL)
        if func_match:
            func_name = func_match.group(1)
            args_str = func_match.group(2)
            return self._explain_function(func_name, args_str)
        
        # 尝试解析二元运算
        # 从右到左找优先级最低的运算符
        op_pos = self._find_lowest_priority_op(formula)
        if op_pos >= 0:
            left = formula[:op_pos].strip()
            right = formula[op_pos+1:].strip()
            op = formula[op_pos]
            
            # 处理双字符运算符
            if op in ('<', '>', '=') and op_pos + 1 < len(formula):
                next_ch = formula[op_pos + 1]
                if next_ch == '=' or (op == '<' and next_ch == '>'):
                    op = op + next_ch
                    right = formula[op_pos+2:].strip()
            
            left_exp = self._parse_and_explain(left)
            right_exp = self._parse_and_explain(right)
            
            op_name = self.OP_NAMES.get(op, op)
            return f"{left_exp} {op_name} {right_exp}"
        
        # 无法解析
        return f"公式: {formula}"
    
    def _find_lowest_priority_op(self, formula: str) -> int:
        """找到最低优先级运算符的位置（从右到左）"""
        depth = 0
        # 优先级从低到高: +-, */, ^
        ops_priority = [('+', 1), ('-', 1), ('*', 2), ('/', 2), ('^', 3), ('&', 4), ('=', 5), ('>', 5), ('<', 5)]
        
        best_pos = -1
        best_priority = 999
        
        for i, ch in enumerate(formula):
            if ch == '(':
                depth += 1
            elif ch == ')':
                depth -= 1
            elif depth == 0:
                # 检查双字符运算符
                two_char = formula[i:i+2]
                if two_char in ('<>', '>=', '<='):
                    op = two_char
                else:
                    op = ch
                
                if op in ['+', '-', '*', '/', '^', '&', '=', '<', '>']:
                    priority = dict(ops_priority).get(op, 999)
                    if priority <= best_priority:
                        best_priority = priority
                        best_pos = i
        
        return best_pos
    
    def _explain_function(self, func_name: str, args_str: str) -> str:
        """解释函数调用"""
        func_upper = func_name.upper()
        func_name_cn = self.FUNC_NAMES.get(func_upper, func_upper)
        
        # 解析参数（考虑嵌套括号）
        args = self._split_args(args_str)
        args_explained = [self._parse_and_explain(a) for a in args]
        
        # 特殊函数特殊处理
        if func_upper == "IF" and len(args_explained) >= 3:
            return f"如果 {args_explained[0]}，则返回 {args_explained[1]}，否则返回 {args_explained[2]}"
        elif func_upper == "IFS" and len(args_explained) >= 2:
            parts = []
            for i in range(0, len(args_explained)-1, 2):
                parts.append(f"如果 {args_explained[i]}，则返回 {args_explained[i+1]}")
            return "；".join(parts)
        elif func_upper == "VLOOKUP" and len(args_explained) >= 4:
            return f"在 {args_explained[1]} 中垂直查找 {args_explained[0]}，返回第 {args_explained[2]} 列的值（精确匹配: {args_explained[3]}）"
        elif func_upper == "HLOOKUP" and len(args_explained) >= 4:
            return f"在 {args_explained[1]} 中水平查找 {args_explained[0]}，返回第 {args_explained[2]} 行的值（精确匹配: {args_explained[3]}）"
        elif func_upper == "INDEX" and len(args_explained) >= 3:
            return f"返回 {args_explained[0]} 的第 {args_explained[1]} 行第 {args_explained[2]} 列的值"
        elif func_upper == "MATCH" and len(args_explained) >= 3:
            return f"在 {args_explained[2]} 中查找 {args_explained[0]} 的位置（匹配类型: {args_explained[1]}）"
        elif func_upper == "COUNTIF" and len(args_explained) >= 2:
            return f"统计 {args_explained[0]} 中满足条件 {args_explained[1]} 的单元格数量"
        elif func_upper == "SUMIF" and len(args_explained) >= 3:
            return f"对 {args_explained[0]} 中满足条件 {args_explained[1]} 的单元格，求 {args_explained[2]} 的和"
        elif func_upper == "IFERROR" and len(args_explained) >= 2:
            return f"如果 {args_explained[0]} 出错，则返回 {args_explained[1]}"
        elif func_upper == "ROUND" and len(args_explained) >= 2:
            return f"将 {args_explained[0]} 四舍五入到 {args_explained[1]} 位小数"
        elif func_upper == "LEFT" and len(args_explained) >= 2:
            return f"从 {args_explained[0]} 左侧取 {args_explained[1]} 个字符"
        elif func_upper == "RIGHT" and len(args_explained) >= 2:
            return f"从 {args_explained[0]} 右侧取 {args_explained[1]} 个字符"
        elif func_upper == "MID" and len(args_explained) >= 3:
            return f"从 {args_explained[0]} 的第 {args_explained[1]} 个字符开始，取 {args_explained[2]} 个字符"
        elif func_upper == "CONCATENATE" or func_upper == "CONCAT":
            return f"连接文本: " + "、".join(args_explained)
        elif func_upper == "AND" and len(args_explained) >= 2:
            return f"所有条件都满足: " + " 且 ".join(args_explained)
        elif func_upper == "OR" and len(args_explained) >= 2:
            return f"任一条件满足: " + " 或 ".join(args_explained)
        elif func_upper == "DATEDIF" and len(args_explained) >= 3:
            return f"计算 {args_explained[0]} 到 {args_explained[1]} 的{args_explained[2]}差"
        elif func_upper == "TEXT" and len(args_explained) >= 2:
            return f"将 {args_explained[0]} 格式化为 {args_explained[1]}"
        
        # 通用函数解释
        args_cn = "、".join(args_explained)
        return f"{func_name_cn}（{args_cn}）"
    
    def _split_args(self, args_str: str) -> List[str]:
        """分割函数参数（考虑嵌套括号）"""
        args = []
        current = ""
        depth = 0
        
        for ch in args_str:
            if ch == '(':
                depth += 1
                current += ch
            elif ch == ')':
                depth -= 1
                current += ch
            elif ch == ',' and depth == 0:
                args.append(current.strip())
                current = ""
            else:
                current += ch
        
        if current.strip():
            args.append(current.strip())
        
        return args
    
    def explain_file(self, filepath: str, sheet: str = "Sheet1") -> List[Dict]:
        """批量解释 Sheet 中所有公式"""
        if not HAS_OPENPYXL:
            return [{"error": "openpyxl 未安装"}]
        
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            if sheet not in wb.sheetnames:
                sheet = wb.sheetnames[0]
            ws = wb[sheet]
            
            formulas = []
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        explanation = self.explain(cell.value)
                        formulas.append({
                            "cell": cell.coordinate,
                            "formula": cell.value,
                            "explanation": explanation,
                        })
            
            wb.close()
            return formulas
        except Exception as e:
            return [{"error": str(e)}]
    
    def explain_cell(self, filepath: str, sheet: str, cell: str) -> Dict:
        """解释单个单元格公式"""
        if not HAS_OPENPYXL:
            return {"error": "openpyxl 未安装"}
        
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            if sheet not in wb.sheetnames:
                sheet = wb.sheetnames[0]
            ws = wb[sheet]
            
            cell_obj = ws[cell]
            if cell_obj.value and isinstance(cell_obj.value, str) and cell_obj.value.startswith("="):
                explanation = self.explain(cell_obj.value)
                result = {
                    "cell": cell,
                    "formula": cell_obj.value,
                    "explanation": explanation,
                }
            else:
                result = {
                    "cell": cell,
                    "formula": None,
                    "explanation": "该单元格不包含公式",
                }
            
            wb.close()
            return result
        except Exception as e:
            return {"error": str(e)}


def _cli():
    """CLI 入口"""
    import argparse
    
    parser = argparse.ArgumentParser(description="公式解释器 v4.7.0")
    sub = parser.add_subparsers(dest="command", required=True)
    
    # explain
    p = sub.add_parser("explain", help="解释单个公式")
    p.add_argument("--formula", required=True, help="Excel 公式（如 =SUM(A1:A10)）")
    
    # explain-file
    p = sub.add_parser("explain-file", help="批量解释 Sheet 中所有公式")
    p.add_argument("--file", required=True, help="Excel 文件路径")
    p.add_argument("--sheet", default="Sheet1")
    
    # explain-cell
    p = sub.add_parser("explain-cell", help="解释单个单元格公式")
    p.add_argument("--file", required=True)
    p.add_argument("--sheet", default="Sheet1")
    p.add_argument("--cell", required=True, help="单元格地址（如 B2）")
    
    args = parser.parse_args()
    
    explainer = FormulaExplainer()
    
    if args.command == "explain":
        print(json.dumps({
            "formula": args.formula,
            "explanation": explainer.explain(args.formula),
        }, ensure_ascii=False))
    
    elif args.command == "explain-file":
        results = explainer.explain_file(args.file, args.sheet)
        print(json.dumps(results, ensure_ascii=False))
    
    elif args.command == "explain-cell":
        result = explainer.explain_cell(args.file, args.sheet, args.cell)
        print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    _cli()
