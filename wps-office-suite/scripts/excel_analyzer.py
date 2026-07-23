"""
Excel 智能分析器 v4.3.0
========================
六大核心能力：
  1. 公式自动纠错（#REF!/#VALUE!/#DIV/0!/#N/A/#NAME?）
  2. 数据清洗辅助（缺失值/重复行/异常值/格式不一致）
  3. 透视表自动生成（智能选行/列/值字段）
  4. 数据可视化建议（基于数据特征推荐图表类型）
  5. 数据预测（移动平均/指数平滑/线性回归）
  6. 自然语言转公式（NL → Excel 公式，规则引擎优先 + LLM 可选降级）

分层架构：
  数据探查层 (DataProfiler) → 智能解析层 (FormulaFixer/PivotRecommender/DataPredictor/NL2Formula) → 输出与可视化层

死规则合规：
  - 规则9：NL2Formula 可接外部 LLM，规则引擎为默认降级
  - 规则10：数据预测根据硬件配置动态降级（低配禁用线性回归）
  - 规则13：不生成任何禁止文件类型
  - 规则14：三轮自审（代码正确性、死规则合规、安全风险、性能、易用性）
  - 规则15：沙箱模拟运行

依赖：openpyxl, pandas, numpy（可选：statsmodels 用于高级预测）
"""
import re
import json
import sys
import os
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
from collections import defaultdict, Counter
from datetime import datetime

# ==================== 硬件自适应 ====================

def get_hardware_level() -> Dict:
    """
    检测硬件配置，返回性能等级
    规则10：低配机器禁用重型计算（线性回归），退化为简单移动平均
    """
    try:
        import multiprocessing
        cpu_count = multiprocessing.cpu_count()
    except Exception:
        cpu_count = 2

    mem_gb = 8  # 默认值
    try:
        if sys.platform == "win32":
            import ctypes
            kernel32 = ctypes.windll.kernel32
            c_ulonglong = ctypes.c_ulonglong

            class MEMORYSTATUSEX(ctypes.Structure):
                _fields_ = [
                    ("dwLength", ctypes.c_ulong),
                    ("dwMemoryLoad", ctypes.c_ulong),
                    ("ullTotalPhys", c_ulonglong),
                    ("ullAvailPhys", c_ulonglong),
                    ("ullTotalPageFile", c_ulonglong),
                    ("ullAvailPageFile", c_ulonglong),
                    ("ullTotalVirtual", c_ulonglong),
                    ("ullAvailVirtual", c_ulonglong),
                    ("ullAvailExtendedVirtual", c_ulonglong),
                ]

            stat = MEMORYSTATUSEX()
            stat.dwLength = ctypes.sizeof(stat)
            kernel32.GlobalMemoryStatusEx(ctypes.byref(stat))
            mem_gb = stat.ullTotalPhys / (1024 ** 3)
        else:
            # Linux / macOS
            try:
                with open("/proc/meminfo", "r") as f:
                    for line in f:
                        if line.startswith("MemTotal"):
                            mem_gb = int(line.split()[1]) / (1024 ** 2)
                            break
            except Exception:
                pass
    except Exception:
        pass

    # 性能分级
    if cpu_count >= 8 and mem_gb >= 16:
        level = "high"
        max_rows_regression = 100000
    elif cpu_count >= 4 and mem_gb >= 8:
        level = "medium"
        max_rows_regression = 50000
    else:
        level = "low"
        max_rows_regression = 0  # 低配禁用线性回归

    return {
        "level": level,
        "cpu_count": cpu_count,
        "mem_gb": round(mem_gb, 1),
        "max_rows_regression": max_rows_regression,
    }


# ==================== 1. 数据探查层 ====================

class DataProfiler:
    """
    数据探查：列类型探测、缺失率统计、唯一值检测、异常值IQR检测、数据质量报告
    """

    def __init__(self, filepath: str, sheet: str = "Sheet1"):
        self.filepath = filepath
        self.sheet = sheet
        self.wb = None
        self.ws = None
        self.headers = []
        self.data = []
        self.n_rows = 0
        self.n_cols = 0
        self._load()

    def _load(self):
        """加载 Excel 数据"""
        try:
            import openpyxl
            self.wb = openpyxl.load_workbook(self.filepath, data_only=True)
            if self.sheet not in self.wb.sheetnames:
                raise ValueError(f"工作表 '{self.sheet}' 不存在，可用: {self.wb.sheetnames}")
            self.ws = self.wb[self.sheet]
            self.headers = [
                str(self.ws.cell(row=1, column=c).value).strip()
                if self.ws.cell(row=1, column=c).value is not None
                else f"列{c}"
                for c in range(1, self.ws.max_column + 1)
            ]
            self.n_cols = len(self.headers)
            self.data = []
            for r in range(2, self.ws.max_row + 1):
                row = []
                for c in range(1, self.n_cols + 1):
                    row.append(self.ws.cell(row=r, column=c).value)
                self.data.append(row)
            self.n_rows = len(self.data)
        except ImportError:
            raise ImportError("需要安装 openpyxl：pip install openpyxl")

    def detect_column_type(self, values: List) -> str:
        """探测单列数据类型：numeric / date / text / empty"""
        non_null = [v for v in values if v is not None and str(v).strip() != ""]
        if not non_null:
            return "empty"

        # 数值检测
        num_count = 0
        for v in non_null:
            try:
                float(v)
                num_count += 1
            except (ValueError, TypeError):
                pass
        if num_count / len(non_null) > 0.8:
            return "numeric"

        # 日期检测
        date_patterns = [
            r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}",
            r"^\d{1,2}[-/]\d{1,2}[-/]\d{4}",
            r"^\d{4}年\d{1,2}月",
        ]
        date_count = 0
        for v in non_null:
            for pat in date_patterns:
                if re.match(pat, str(v).strip()):
                    date_count += 1
                    break
        if date_count / len(non_null) > 0.7:
            return "date"

        return "text"

    def get_column_values(self, col_idx: int) -> List:
        """获取列值（跳过表头）"""
        return [row[col_idx] for row in self.data if col_idx < len(row)]

    def profile(self) -> Dict:
        """生成完整数据质量报告"""
        column_profiles = []
        total_cells = self.n_rows * self.n_cols
        total_missing = 0
        total_duplicates = 0

        for col_idx in range(self.n_cols):
            values = self.get_column_values(col_idx)
            col_type = self.detect_column_type(values)
            non_null = [v for v in values if v is not None and str(v).strip() != ""]
            missing_count = len(values) - len(non_null)
            total_missing += missing_count

            # 唯一值统计
            unique_values = set(str(v) for v in non_null)
            unique_count = len(unique_values)

            # 异常值检测（仅数值列）
            outliers = []
            if col_type == "numeric" and len(non_null) >= 4:
                nums = []
                for v in non_null:
                    try:
                        nums.append(float(v))
                    except (ValueError, TypeError):
                        pass
                if len(nums) >= 4:
                    nums_sorted = sorted(nums)
                    q1_idx = len(nums_sorted) // 4
                    q3_idx = 3 * len(nums_sorted) // 4
                    q1 = nums_sorted[q1_idx]
                    q3 = nums_sorted[q3_idx]
                    iqr = q3 - q1
                    lower = q1 - 1.5 * iqr
                    upper = q3 + 1.5 * iqr
                    outliers = [v for v in nums if v < lower or v > upper]

            column_profiles.append({
                "name": self.headers[col_idx],
                "index": col_idx,
                "type": col_type,
                "non_null_count": len(non_null),
                "missing_count": missing_count,
                "missing_rate": round(missing_count / max(len(values), 1), 4),
                "unique_count": unique_count,
                "unique_rate": round(unique_count / max(len(non_null), 1), 4),
                "outlier_count": len(outliers),
                "outlier_values": outliers[:10],  # 最多返回10个异常值
                "sample_values": non_null[:5],
            })

        # 重复行检测
        seen = set()
        for row in self.data:
            row_key = tuple(str(c) if c is not None else "" for c in row)
            if row_key in seen:
                total_duplicates += 1
            seen.add(row_key)

        # 整体质量评分
        quality_score = 100
        if total_cells > 0:
            quality_score -= int((total_missing / total_cells) * 30)
            quality_score -= int((total_duplicates / max(self.n_rows, 1)) * 20)
        quality_score = max(0, min(100, quality_score))

        return {
            "success": True,
            "file": self.filepath,
            "sheet": self.sheet,
            "n_rows": self.n_rows,
            "n_cols": self.n_cols,
            "total_cells": total_cells,
            "total_missing": total_missing,
            "total_duplicates": total_duplicates,
            "missing_rate": round(total_missing / max(total_cells, 1), 4),
            "duplicate_rate": round(total_duplicates / max(self.n_rows, 1), 4),
            "quality_score": quality_score,
            "columns": column_profiles,
            "numeric_columns": [c["name"] for c in column_profiles if c["type"] == "numeric"],
            "text_columns": [c["name"] for c in column_profiles if c["type"] == "text"],
            "date_columns": [c["name"] for c in column_profiles if c["type"] == "date"],
        }


# ==================== 2. 公式自动纠错 ====================

class FormulaFixer:
    """
    公式纠错引擎：规则引擎修复常见 Excel 错误
    优先级：规则引擎（本地）> 外部 LLM（可选，需降级）
    """

    # 常见错误模式及修复策略
    ERROR_PATTERNS = {
        "#REF!": {
            "description": "引用无效单元格",
            "fix_type": "rewrite_ref",
            "suggestion": "检查引用范围是否正确，可能删除了被引用的行/列",
        },
        "#VALUE!": {
            "description": "数据类型不匹配",
            "fix_type": "type_convert",
            "suggestion": "确保参与运算的单元格为数值类型，可用 VALUE() 转换",
        },
        "#DIV/0!": {
            "description": "除以零",
            "fix_type": "iferror_wrap",
            "suggestion": "用 IFERROR(公式, 0) 包裹，避免除以零",
        },
        "#N/A": {
            "description": "查找函数未找到匹配值",
            "fix_type": "ifna_wrap",
            "suggestion": "用 IFNA(公式, \"未找到\") 包裹，或检查查找值是否存在",
        },
        "#NAME?": {
            "description": "函数名拼写错误或未定义名称",
            "fix_type": "fix_name",
            "suggestion": "检查函数名拼写，确保已加载对应加载项",
        },
        "#NULL!": {
            "description": "区域交集运算符使用错误",
            "fix_type": "fix_range",
            "suggestion": "检查区域引用，确保使用了正确的交集运算符",
        },
        "#NUM!": {
            "description": "数值计算错误",
            "fix_type": "fix_numeric",
            "suggestion": "检查数值范围，避免无效的数学运算（如负数的平方根）",
        },
    }

    # 常见拼写错误映射
    COMMON_MISSPELLINGS = {
        "SUMM": "SUM",
        "AVERGE": "AVERAGE",
        "AVG": "AVERAGE",
        "COUTN": "COUNT",
        "CONCAT": "CONCATENATE",
        "VLOKUP": "VLOOKUP",
        "VLOOCKUP": "VLOOKUP",
        "HLOKUP": "HLOOKUP",
        "SUMMIF": "SUMIF",
        "COUNTIFS": "COUNTIFS",
        "SUMIFS": "SUMIFS",
        "IFEROR": "IFERROR",
        "IFERORR": "IFERROR",
    }

    def __init__(self, filepath: str, sheet: str = "Sheet1"):
        self.filepath = filepath
        self.sheet = sheet
        self.wb = None
        self.ws = None
        self._load()

    def _load(self):
        try:
            import openpyxl
            self.wb = openpyxl.load_workbook(self.filepath)
            if self.sheet not in self.wb.sheetnames:
                raise ValueError(f"工作表 '{self.sheet}' 不存在")
            self.ws = self.wb[self.sheet]
        except ImportError:
            raise ImportError("需要安装 openpyxl：pip install openpyxl")

    def scan_errors(self) -> List[Dict]:
        """扫描工作表中的所有错误"""
        errors = []
        for row_idx in range(1, self.ws.max_row + 1):
            for col_idx in range(1, self.ws.max_column + 1):
                cell = self.ws.cell(row=row_idx, column=col_idx)
                value = cell.value
                if value is None:
                    continue
                value_str = str(value)
                for err_code, err_info in self.ERROR_PATTERNS.items():
                    if err_code in value_str:
                        errors.append({
                            "cell": cell.coordinate,
                            "row": row_idx,
                            "col": col_idx,
                            "error": err_code,
                            "description": err_info["description"],
                            "fix_type": err_info["fix_type"],
                            "suggestion": err_info["suggestion"],
                            "original_formula": value_str,
                        })
        return errors

    def fix_formula(self, formula: str, error_type: str) -> Optional[str]:
        """
        根据错误类型自动修复公式
        返回修复后的公式，或 None 如果无法自动修复
        """
        if error_type == "#DIV/0!":
            # 用 IFERROR 包裹
            if not formula.upper().startswith("IFERROR"):
                return f"IFERROR({formula}, 0)"
        elif error_type == "#N/A":
            # 用 IFNA 包裹
            if not formula.upper().startswith("IFNA"):
                return f"IFNA({formula}, \"未找到\")"
        elif error_type == "#NAME?":
            # 修复拼写错误
            fixed = formula
            for wrong, correct in self.COMMON_MISSPELLINGS.items():
                fixed = re.sub(r'\b' + re.escape(wrong) + r'\b', correct, fixed, flags=re.IGNORECASE)
            if fixed != formula:
                return fixed
        elif error_type == "#VALUE!":
            # 尝试用 VALUE() 包裹文本型数字
            # 简单处理：如果公式中有文本运算，尝试转换
            return f"IFERROR(VALUE({formula}), 0)"
        return None

    def fix_all(self, use_llm: bool = False) -> Dict:
        """
        扫描并修复所有错误
        use_llm: 是否使用外部 LLM 辅助修复复杂错误（规则9：可选，有降级）
        """
        errors = self.scan_errors()
        fixed = []
        failed = []
        skipped = []

        for err in errors:
            formula = err["original_formula"]
            fix_type = err["fix_type"]

            # 规则引擎修复
            new_formula = self.fix_formula(formula, err["error"])

            if new_formula:
                # 应用修复
                cell = self.ws.cell(row=err["row"], column=err["col"])
                cell.value = new_formula
                fixed.append({
                    "cell": err["cell"],
                    "original": formula,
                    "fixed": new_formula,
                    "method": "rule_engine",
                })
            else:
                # 规则引擎无法修复，尝试 LLM（可选）
                if use_llm:
                    llm_result = self._try_llm_fix(formula, err["error"])
                    if llm_result:
                        cell = self.ws.cell(row=err["row"], column=err["col"])
                        cell.value = llm_result
                        fixed.append({
                            "cell": err["cell"],
                            "original": formula,
                            "fixed": llm_result,
                            "method": "llm_assist",
                        })
                    else:
                        failed.append({
                            "cell": err["cell"],
                            "error": err["error"],
                            "reason": "LLM 也无法修复",
                        })
                else:
                    skipped.append({
                        "cell": err["cell"],
                        "error": err["error"],
                        "suggestion": err["suggestion"],
                    })

        # 保存修复后的文件
        if fixed:
            self.wb.save(self.filepath)

        return {
            "success": True,
            "total_errors": len(errors),
            "fixed_count": len(fixed),
            "failed_count": len(failed),
            "skipped_count": len(skipped),
            "fixed": fixed,
            "failed": failed,
            "skipped": skipped,
        }

    def _try_llm_fix(self, formula: str, error_type: str) -> Optional[str]:
        """
        尝试使用外部 LLM 修复公式（规则9：可选，有降级）
        当 LLM 不可用时返回 None，由调用方处理
        """
        # 这里预留 LLM 接口，实际调用时需配置 API
        # 降级方案：返回 None，由规则引擎的 suggestion 提供人工修复建议
        return None


# ==================== 3. 透视表自动生成 ====================

class PivotRecommender:
    """
    透视表推荐引擎：根据数据特征自动推荐行/列/值字段和聚合方式
    """

    def __init__(self, filepath: str, sheet: str = "Sheet1"):
        self.filepath = filepath
        self.sheet = sheet
        self.profiler = DataProfiler(filepath, sheet)

    def recommend(self) -> Dict:
        """生成透视表推荐配置"""
        profile = self.profiler.profile()
        if not profile.get("success"):
            return profile

        numeric_cols = profile["numeric_columns"]
        text_cols = profile["text_columns"]
        date_cols = profile["date_columns"]

        if not numeric_cols:
            return {
                "success": False,
                "error": "没有数值型列，无法生成透视表",
                "suggestion": "透视表需要至少一个数值列作为值字段",
            }

        if not text_cols and not date_cols:
            return {
                "success": False,
                "error": "没有文本或日期列作为分组字段",
                "suggestion": "透视表需要至少一个分类字段作为行/列",
            }

        # 智能选择行字段：优先文本列（唯一值较少的）
        row_field = None
        for col in text_cols:
            col_info = next((c for c in profile["columns"] if c["name"] == col), None)
            if col_info and 1 < col_info["unique_count"] <= min(self.profiler.n_rows // 2, 50):
                row_field = col
                break
        if not row_field and text_cols:
            row_field = text_cols[0]
        if not row_field and date_cols:
            row_field = date_cols[0]

        # 列字段：选择第二个分类字段（可选）
        col_field = None
        remaining = [c for c in text_cols if c != row_field]
        if remaining:
            col_field = remaining[0]

        # 值字段：选择第一个数值列
        value_field = numeric_cols[0]

        # 聚合方式：根据数据特征选择
        agg_func = "sum"
        if len(numeric_cols) > 1:
            # 如果有多个数值列，检查是否有比率类字段
            for nc in numeric_cols:
                if any(kw in nc.lower() for kw in ["率", "比例", "占比", "rate", "ratio", "%"]):
                    value_field = nc
                    agg_func = "avg"
                    break

        return {
            "success": True,
            "recommendation": {
                "row_field": row_field,
                "col_field": col_field,
                "value_field": value_field,
                "agg_func": agg_func,
                "explanation": f"按「{row_field}」分组，对「{value_field}」求{agg_func}",
            },
            "alternatives": [
                {
                    "row_field": row_field,
                    "value_field": vf,
                    "agg_func": "sum",
                }
                for vf in numeric_cols[:3] if vf != value_field
            ],
        }

    def generate(self, row_field: str, value_field: str,
                 col_field: str = None, agg_func: str = "sum") -> Dict:
        """
        生成透视表并写入新 Sheet
        """
        try:
            import openpyxl

            wb = openpyxl.load_workbook(self.filepath)
            ws = wb[self.sheet]

            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            headers = [str(h).strip() if h else f"列{i+1}" for i, h in enumerate(headers)]

            # 找到字段索引
            try:
                row_idx = headers.index(row_field)
            except ValueError:
                return {"success": False, "error": f"行字段 '{row_field}' 不存在"}
            try:
                val_idx = headers.index(value_field)
            except ValueError:
                return {"success": False, "error": f"值字段 '{value_field}' 不存在"}

            col_idx = None
            if col_field:
                try:
                    col_idx = headers.index(col_field)
                except ValueError:
                    col_field = None

            # 读取数据
            data = []
            for r in range(2, ws.max_row + 1):
                row = []
                for c in range(1, len(headers) + 1):
                    row.append(ws.cell(row=r, column=c).value)
                data.append(row)

            # 分组聚合
            grouped = defaultdict(list)
            for row in data:
                if len(row) <= max(row_idx, val_idx):
                    continue
                key = str(row[row_idx]) if row[row_idx] is not None else "未知"
                if col_field and col_idx is not None:
                    col_key = str(row[col_idx]) if row[col_idx] is not None else "未知"
                    key = (key, col_key)
                try:
                    val = float(row[val_idx]) if row[val_idx] is not None else 0
                    grouped[key].append(val)
                except (ValueError, TypeError):
                    pass

            # 计算聚合
            results = []
            for key, values in grouped.items():
                if agg_func == "sum":
                    result = sum(values)
                elif agg_func == "avg":
                    result = sum(values) / len(values) if values else 0
                elif agg_func == "count":
                    result = len(values)
                elif agg_func == "max":
                    result = max(values) if values else 0
                elif agg_func == "min":
                    result = min(values) if values else 0
                else:
                    result = sum(values)

                results.append({
                    "key": key,
                    "value": round(result, 2),
                    "count": len(values),
                })

            results.sort(key=lambda x: x["value"], reverse=True)

            # 写入新 Sheet
            pivot_sheet_name = f"透视_{row_field}"
            if pivot_sheet_name in wb.sheetnames:
                del wb[pivot_sheet_name]
            pivot_ws = wb.create_sheet(pivot_sheet_name)

            # 写入表头
            pivot_ws.cell(row=1, column=1, value=row_field)
            if col_field:
                pivot_ws.cell(row=1, column=2, value=col_field)
            pivot_ws.cell(row=1, column=2 if not col_field else 3, value=f"{agg_func}({value_field})")
            pivot_ws.cell(row=1, column=3 if not col_field else 4, value="计数")

            # 写入数据
            for i, r in enumerate(results):
                row_num = i + 2
                if col_field and isinstance(r["key"], tuple):
                    pivot_ws.cell(row=row_num, column=1, value=r["key"][0])
                    pivot_ws.cell(row=row_num, column=2, value=r["key"][1])
                    pivot_ws.cell(row=row_num, column=3, value=r["value"])
                    pivot_ws.cell(row=row_num, column=4, value=r["count"])
                else:
                    pivot_ws.cell(row=row_num, column=1, value=str(r["key"]))
                    pivot_ws.cell(row=row_num, column=2, value=r["value"])
                    pivot_ws.cell(row=row_num, column=3, value=r["count"])

            wb.save(self.filepath)

            return {
                "success": True,
                "sheet_name": pivot_sheet_name,
                "row_count": len(results),
                "row_field": row_field,
                "value_field": value_field,
                "agg_func": agg_func,
                "results": results[:50],  # 最多返回50行
            }

        except ImportError:
            return {"success": False, "error": "需要安装 openpyxl"}
        except Exception as e:
            return {"success": False, "error": str(e)}


# ==================== 4. 数据预测 ====================

class DataPredictor:
    """
    数据预测引擎：移动平均、指数平滑、线性回归
    规则10：低配硬件禁用线性回归，退化为简单移动平均
    """

    def __init__(self, filepath: str, sheet: str = "Sheet1"):
        self.filepath = filepath
        self.sheet = sheet
        self.hardware = get_hardware_level()

    def moving_average(self, values: List[float], window: int = 3) -> List[float]:
        """简单移动平均"""
        result = []
        for i in range(len(values)):
            if i < window - 1:
                result.append(None)
            else:
                avg = sum(values[i - window + 1:i + 1]) / window
                result.append(round(avg, 4))
        return result

    def exponential_smoothing(self, values: List[float], alpha: float = 0.3) -> List[float]:
        """指数平滑"""
        if not values:
            return []
        result = [values[0]]
        for i in range(1, len(values)):
            smoothed = alpha * values[i] + (1 - alpha) * result[-1]
            result.append(round(smoothed, 4))
        return result

    def linear_regression(self, values: List[float], forecast_steps: int = 3) -> Optional[Dict]:
        """
        线性回归预测
        规则10：低配硬件（max_rows_regression=0）禁用此方法
        """
        if self.hardware["max_rows_regression"] == 0:
            return None  # 低配禁用

        try:
            import numpy as np
        except ImportError:
            return None

        n = len(values)
        if n < 3:
            return None

        x = np.arange(n)
        y = np.array(values, dtype=float)

        # 最小二乘法
        x_mean = x.mean()
        y_mean = y.mean()
        ss_xy = np.sum((x - x_mean) * (y - y_mean))
        ss_xx = np.sum((x - x_mean) ** 2)

        if ss_xx == 0:
            return None

        slope = ss_xy / ss_xx
        intercept = y_mean - slope * x_mean

        # 拟合值
        fitted = slope * x + intercept

        # 预测
        future_x = np.arange(n, n + forecast_steps)
        forecast = slope * future_x + intercept

        # R²
        ss_res = np.sum((y - fitted) ** 2)
        ss_tot = np.sum((y - y_mean) ** 2)
        r_squared = 1 - ss_res / ss_tot if ss_tot != 0 else 0

        return {
            "slope": round(float(slope), 6),
            "intercept": round(float(intercept), 6),
            "r_squared": round(float(r_squared), 4),
            "fitted_values": [round(float(v), 4) for v in fitted],
            "forecast_values": [round(float(v), 4) for v in forecast],
            "forecast_steps": forecast_steps,
        }

    def predict(self, column_name: str, method: str = "auto",
                forecast_steps: int = 3) -> Dict:
        """
        对指定列进行预测
        method: auto / moving_avg / exponential / linear
        """
        try:
            import openpyxl

            wb = openpyxl.load_workbook(self.filepath, data_only=True)
            if self.sheet not in wb.sheetnames:
                return {"success": False, "error": f"工作表 '{self.sheet}' 不存在"}

            ws = wb[self.sheet]
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            headers = [str(h).strip() if h else f"列{i+1}" for i, h in enumerate(headers)]

            if column_name not in headers:
                return {"success": False, "error": f"列 '{column_name}' 不存在"}

            col_idx = headers.index(column_name)
            values = []
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=col_idx + 1).value
                try:
                    values.append(float(v))
                except (ValueError, TypeError):
                    values.append(0.0)

            if len(values) < 3:
                return {"success": False, "error": "数据点太少（至少需要3个）"}

            # 自动选择方法
            if method == "auto":
                if len(values) >= 10 and self.hardware["max_rows_regression"] > 0:
                    method = "linear"
                elif len(values) >= 5:
                    method = "exponential"
                else:
                    method = "moving_avg"

            result = {
                "success": True,
                "column": column_name,
                "method": method,
                "data_points": len(values),
                "original_values": values,
            }

            if method == "moving_avg":
                ma = self.moving_average(values, window=min(3, len(values) // 2 + 1))
                result["smoothed_values"] = ma
                # 简单预测：用最后一个移动平均值
                last_ma = [v for v in ma if v is not None]
                if last_ma:
                    result["forecast_values"] = [round(last_ma[-1], 2)] * forecast_steps

            elif method == "exponential":
                smoothed = self.exponential_smoothing(values)
                result["smoothed_values"] = smoothed
                result["forecast_values"] = [round(smoothed[-1], 2)] * forecast_steps

            elif method == "linear":
                lr = self.linear_regression(values, forecast_steps)
                if lr:
                    result.update(lr)
                    result["method"] = "linear"
                else:
                    # 降级为指数平滑
                    smoothed = self.exponential_smoothing(values)
                    result["smoothed_values"] = smoothed
                    result["forecast_values"] = [round(smoothed[-1], 2)] * forecast_steps
                    result["method"] = "exponential (降级)"
                    result["degraded_reason"] = "硬件配置不足或数据点不够，线性回归已降级"

            return result

        except ImportError:
            return {"success": False, "error": "需要安装 openpyxl"}
        except Exception as e:
            return {"success": False, "error": str(e)}


# ==================== 5. 自然语言转公式 ====================

class NL2Formula:
    """
    自然语言转 Excel 公式
    规则9：规则引擎优先，外部 LLM 可选 + 降级
    """

    # 公式模板库
    FORMULA_TEMPLATES = [
        {
            "patterns": ["同比增长率", "同比增长", "增长率", "同比"],
            "formula": "=(本期值-同期值)/同期值",
            "explanation": "同比增长率 = (本期值 - 同期值) / 同期值",
            "example": "=(B3-B2)/B2",
        },
        {
            "patterns": ["环比增长率", "环比增长", "环比"],
            "formula": "=(本期值-上期值)/上期值",
            "explanation": "环比增长率 = (本期值 - 上期值) / 上期值",
            "example": "=(B3-B2)/B2",
        },
        {
            "patterns": ["占比", "比例", "百分比", "份额"],
            "formula": "=部分值/总值",
            "explanation": "占比 = 部分值 / 总值",
            "example": "=B2/SUM(B:B)",
        },
        {
            "patterns": ["累计", "累计求和", "累计和"],
            "formula": "=SUM(起始:当前)",
            "explanation": "累计求和 = SUM(起始行:当前行)",
            "example": "=SUM(B$2:B2)",
        },
        {
            "patterns": ["排名", "排序", "名次"],
            "formula": "=RANK(值, 区域)",
            "explanation": "排名 = RANK(值, 排序区域)",
            "example": "=RANK(B2,B:B)",
        },
        {
            "patterns": ["平均值", "平均", "均值"],
            "formula": "=AVERAGE(区域)",
            "explanation": "平均值 = AVERAGE(数据区域)",
            "example": "=AVERAGE(B2:B10)",
        },
        {
            "patterns": ["最大值", "最高", "最大"],
            "formula": "=MAX(区域)",
            "explanation": "最大值 = MAX(数据区域)",
            "example": "=MAX(B2:B10)",
        },
        {
            "patterns": ["最小值", "最低", "最小"],
            "formula": "=MIN(区域)",
            "explanation": "最小值 = MIN(数据区域)",
            "example": "=MIN(B2:B10)",
        },
        {
            "patterns": ["计数", "数量", "个数"],
            "formula": "=COUNTA(区域)",
            "explanation": "计数 = COUNTA(数据区域)",
            "example": "=COUNTA(A2:A10)",
        },
        {
            "patterns": ["条件求和", "按条件求和"],
            "formula": "=SUMIF(条件区域, 条件, 求和区域)",
            "explanation": "条件求和 = SUMIF(条件区域, 条件, 求和区域)",
            "example": "=SUMIF(A:A,\"华东\",B:B)",
        },
        {
            "patterns": ["条件计数", "按条件计数"],
            "formula": "=COUNTIF(条件区域, 条件)",
            "explanation": "条件计数 = COUNTIF(条件区域, 条件)",
            "example": "=COUNTIF(A:A,\"华东\")",
        },
        {
            "patterns": ["查找", "匹配", "vlookup"],
            "formula": "=VLOOKUP(查找值, 区域, 列号, FALSE)",
            "explanation": "查找 = VLOOKUP(查找值, 数据区域, 返回列号, 精确匹配)",
            "example": "=VLOOKUP(D2,A:B,2,FALSE)",
        },
        {
            "patterns": ["四舍五入", "保留小数", "round"],
            "formula": "=ROUND(值, 小数位数)",
            "explanation": "四舍五入 = ROUND(数值, 保留小数位数)",
            "example": "=ROUND(B2,2)",
        },
        {
            "patterns": ["if判断", "如果", "条件"],
            "formula": "=IF(条件, 真值, 假值)",
            "explanation": "条件判断 = IF(条件, 条件为真时的值, 条件为假时的值)",
            "example": "=IF(B2>100,\"高\",\"低\")",
        },
    ]

    def __init__(self, filepath: str = None, sheet: str = "Sheet1"):
        self.filepath = filepath
        self.sheet = sheet

    def convert(self, query: str, use_llm: bool = False) -> Dict:
        """
        将自然语言转换为 Excel 公式
        规则9：规则引擎优先，LLM 可选 + 降级
        """
        # 规则引擎匹配
        best_match = None
        best_score = 0

        for template in self.FORMULA_TEMPLATES:
            for pattern in template["patterns"]:
                if pattern in query:
                    score = len(pattern)
                    if score > best_score:
                        best_score = score
                        best_match = template
                    break

        if best_match:
            return {
                "success": True,
                "query": query,
                "formula": best_match["formula"],
                "explanation": best_match["explanation"],
                "example": best_match["example"],
                "method": "rule_engine",
            }

        # 规则引擎无法匹配，尝试 LLM（可选）
        if use_llm:
            llm_result = self._try_llm_convert(query)
            if llm_result:
                return llm_result

        # 降级：返回建议
        return {
            "success": False,
            "query": query,
            "error": "无法自动识别公式意图",
            "suggestion": "请尝试更明确的描述，如：求和、平均值、同比增长率、占比、排名等",
            "available_patterns": [
                "求和", "平均值", "最大值", "最小值", "计数",
                "同比增长率", "环比增长率", "占比", "排名",
                "条件求和", "条件计数", "查找", "四舍五入", "if判断",
            ],
        }

    def _try_llm_convert(self, query: str) -> Optional[Dict]:
        """
        尝试使用外部 LLM 转换（规则9：可选，有降级）
        """
        # 预留 LLM 接口
        return None


# ==================== 6. 数据清洗辅助 ====================

class DataCleaner:
    """
    数据清洗辅助：识别缺失值、重复行、异常值、格式不一致
    """

    def __init__(self, filepath: str, sheet: str = "Sheet1"):
        self.filepath = filepath
        self.sheet = sheet
        self.profiler = DataProfiler(filepath, sheet)

    def analyze_cleaning(self) -> Dict:
        """分析数据清洗需求"""
        profile = self.profiler.profile()
        if not profile.get("success"):
            return profile

        issues = []
        suggestions = []

        # 缺失值检测
        high_missing_cols = [
            c for c in profile["columns"]
            if c["missing_rate"] > 0.1 and c["missing_rate"] < 1.0
        ]
        if high_missing_cols:
            issues.append({
                "type": "missing_values",
                "severity": "medium",
                "description": f"发现 {len(high_missing_cols)} 个列有缺失值",
                "columns": [{"name": c["name"], "missing_rate": c["missing_rate"]} for c in high_missing_cols],
            })
            suggestions.append("建议：对缺失值进行填充（均值/中位数/众数）或删除")

        # 重复行检测
        if profile["total_duplicates"] > 0:
            issues.append({
                "type": "duplicate_rows",
                "severity": "high",
                "description": f"发现 {profile['total_duplicates']} 行重复数据",
            })
            suggestions.append("建议：删除重复行，保留唯一记录")

        # 异常值检测
        outlier_cols = [c for c in profile["columns"] if c["outlier_count"] > 0]
        if outlier_cols:
            issues.append({
                "type": "outliers",
                "severity": "low",
                "description": f"发现 {len(outlier_cols)} 个列存在异常值",
                "columns": [{"name": c["name"], "outlier_count": c["outlier_count"]} for c in outlier_cols],
            })
            suggestions.append("建议：检查异常值是否为录入错误，或进行截断处理")

        # 格式不一致检测（文本列中的数字）
        format_issues = []
        for c in profile["columns"]:
            if c["type"] == "text":
                # 检查是否混有数字
                text_values = [str(v) for v in c.get("sample_values", []) if v is not None]
                has_number = any(re.match(r'^-?\d+\.?\d*$', str(v)) for v in text_values)
                if has_number:
                    format_issues.append(c["name"])
        if format_issues:
            issues.append({
                "type": "format_inconsistency",
                "severity": "medium",
                "description": f"文本列中混有数值: {', '.join(format_issues)}",
            })
            suggestions.append("建议：统一数据类型，将数值字符串转为数值")

        return {
            "success": True,
            "quality_score": profile["quality_score"],
            "total_issues": len(issues),
            "issues": issues,
            "suggestions": suggestions,
            "profile_summary": {
                "n_rows": profile["n_rows"],
                "n_cols": profile["n_cols"],
                "missing_rate": profile["missing_rate"],
                "duplicate_rate": profile["duplicate_rate"],
            },
        }


# ==================== 统一入口 ====================

def analyze_excel(filepath: str, task: str, **kwargs) -> Dict:
    """
    Excel 智能分析统一入口

    Args:
        filepath: Excel 文件路径
        task: 任务类型
          - profile: 数据探查
          - fix_formulas: 公式纠错
          - pivot: 透视表生成
          - predict: 数据预测
          - nl2formula: 自然语言转公式
          - clean: 数据清洗分析
        **kwargs: 额外参数

    Returns:
        Dict: 分析结果
    """
    if not Path(filepath).exists():
        return {"success": False, "error": f"文件不存在: {filepath}"}

    sheet = kwargs.get("sheet", "Sheet1")

    try:
        if task == "profile":
            profiler = DataProfiler(filepath, sheet)
            return profiler.profile()

        elif task == "fix_formulas":
            fixer = FormulaFixer(filepath, sheet)
            use_llm = kwargs.get("use_llm", False)
            return fixer.fix_all(use_llm=use_llm)

        elif task == "pivot":
            recommender = PivotRecommender(filepath, sheet)
            row_field = kwargs.get("row_field")
            value_field = kwargs.get("value_field")
            if row_field and value_field:
                return recommender.generate(
                    row_field=row_field,
                    value_field=value_field,
                    col_field=kwargs.get("col_field"),
                    agg_func=kwargs.get("agg_func", "sum"),
                )
            else:
                return recommender.recommend()

        elif task == "predict":
            predictor = DataPredictor(filepath, sheet)
            return predictor.predict(
                column_name=kwargs["column_name"],
                method=kwargs.get("method", "auto"),
                forecast_steps=kwargs.get("forecast_steps", 3),
            )

        elif task == "nl2formula":
            converter = NL2Formula(filepath, sheet)
            return converter.convert(
                query=kwargs["query"],
                use_llm=kwargs.get("use_llm", False),
            )

        elif task == "clean":
            cleaner = DataCleaner(filepath, sheet)
            return cleaner.analyze_cleaning()

        else:
            return {
                "success": False,
                "error": f"未知任务类型: {task}",
                "available_tasks": ["profile", "fix_formulas", "pivot", "predict", "nl2formula", "clean"],
            }

    except Exception as e:
        return {"success": False, "error": str(e)}


# ==================== CLI 入口 ====================

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Excel 智能分析器 v4.3.0")
    sub = parser.add_subparsers(dest="command", required=True)

    # profile
    p_profile = sub.add_parser("profile", help="数据探查")
    p_profile.add_argument("--file", required=True)
    p_profile.add_argument("--sheet", default="Sheet1")

    # fix-formulas
    p_fix = sub.add_parser("fix-formulas", help="公式纠错")
    p_fix.add_argument("--file", required=True)
    p_fix.add_argument("--sheet", default="Sheet1")
    p_fix.add_argument("--use-llm", action="store_true", help="使用 LLM 辅助修复（可选）")

    # pivot
    p_pivot = sub.add_parser("pivot", help="透视表生成")
    p_pivot.add_argument("--file", required=True)
    p_pivot.add_argument("--sheet", default="Sheet1")
    p_pivot.add_argument("--row-field", help="行字段（不指定则自动推荐）")
    p_pivot.add_argument("--value-field", help="值字段")
    p_pivot.add_argument("--col-field", help="列字段（可选）")
    p_pivot.add_argument("--agg-func", default="sum", help="聚合方式")

    # predict
    p_predict = sub.add_parser("predict", help="数据预测")
    p_predict.add_argument("--file", required=True)
    p_predict.add_argument("--sheet", default="Sheet1")
    p_predict.add_argument("--column", required=True, help="预测列名")
    p_predict.add_argument("--method", default="auto", help="预测方法: auto/moving_avg/exponential/linear")
    p_predict.add_argument("--steps", type=int, default=3, help="预测步数")

    # nl2formula
    p_nl2f = sub.add_parser("nl2formula", help="自然语言转公式")
    p_nl2f.add_argument("--query", required=True, help="自然语言描述")
    p_nl2f.add_argument("--file", help="Excel 文件路径（可选，用于上下文）")
    p_nl2f.add_argument("--sheet", default="Sheet1")
    p_nl2f.add_argument("--use-llm", action="store_true")

    # clean
    p_clean = sub.add_parser("clean", help="数据清洗分析")
    p_clean.add_argument("--file", required=True)
    p_clean.add_argument("--sheet", default="Sheet1")

    # hardware
    p_hw = sub.add_parser("hardware", help="查看硬件配置")

    args = parser.parse_args()

    if args.command == "profile":
        result = analyze_excel(args.file, "profile", sheet=args.sheet)
    elif args.command == "fix-formulas":
        result = analyze_excel(args.file, "fix_formulas", sheet=args.sheet, use_llm=args.use_llm)
    elif args.command == "pivot":
        result = analyze_excel(
            args.file, "pivot", sheet=args.sheet,
            row_field=args.row_field, value_field=args.value_field,
            col_field=args.col_field, agg_func=args.agg_func,
        )
    elif args.command == "predict":
        result = analyze_excel(
            args.file, "predict", sheet=args.sheet,
            column_name=args.column, method=args.method, forecast_steps=args.steps,
        )
    elif args.command == "nl2formula":
        result = analyze_excel(
            args.file or "", "nl2formula",
            query=args.query, use_llm=args.use_llm,
        )
    elif args.command == "clean":
        result = analyze_excel(args.file, "clean", sheet=args.sheet)
    elif args.command == "hardware":
        result = get_hardware_level()
    else:
        result = {"success": False, "error": "未知命令"}

    print(json.dumps(result, ensure_ascii=False, default=str, indent=2))
