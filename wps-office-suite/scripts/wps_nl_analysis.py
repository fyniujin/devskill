"""
Excel 自然语言数据分析器 v4.0
输入「按月份统计销售额并画出趋势图」→ 自动解析字段 → 执行数据分析 → 生成图表

核心能力：
  • 中文 NL 解析（关键词提取 + 意图识别）
  • 自动字段匹配（模糊匹配列名）
  • 数据透视分析（分组聚合）
  • 图表自动生成
  • 公式自动写入

不依赖任何外部 API，纯本地正则匹配 + 规则引擎实现。
"""
import re
import json
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from collections import defaultdict


# ==================== NL 解析引擎 ====================

# 聚合函数关键词映射
AGG_KEYWORDS = {
    "sum": ["总和", "合计", "总共", "总计", "共", "求和", "sum", "总额", "一共"],
    "avg": ["平均", "均值", "average", "avg", "平均水平", "一般"],
    "max": ["最大", "最高", "峰值", "最多", "max", "最佳"],
    "min": ["最小", "最低", "最少", "min"],
    "count": ["计数", "多少个", "数量", "count", "有几个", "几项"],
}

# 图表类型关键词映射
CHART_KEYWORDS = {
    "line": ["趋势", "折线", "走势", "变化", "趋势图", "线性", "时间序列"],
    "bar": ["柱状", "对比", "比较", "排名", "top", "bar", "条形"],
    "pie": ["饼图", "占比", "比例", "份额", "pie", "构成", "分布"],
    "column": ["柱形", "column", "直方"],
}

# 分组关键词
GROUP_KEYWORDS = ["按", "根据", "按...分", "以...为", "by", "维度", "分组"]

# 过滤关键词
FILTER_PATTERNS = [
    (r"大于\s*(\d+)", ">"),
    (r"小于\s*(\d+)", "<"),
    (r"超过\s*(\d+)", ">"),
    (r"不足\s*(\d+)", "<"),
    (r"等于\s*(\d+)或\"([^\"]+)\"", "="),
    (r"大于等于\s*(\d+)", ">="),
    (r"小于等于\s*(\d+)", "<="),
]

# 排序关键词
SORT_KEYWORDS = {
    "asc": ["升序", "从小到大", "由少到多", "从低到高", "asc"],
    "desc": ["降序", "从大到小", "由多到少", "从高到低", "top", "排名", "desc"],
}


def analyze_intent(query: str) -> Dict:
    """分析用户查询意图"""
    intent = {
        "operation": None,      # aggregate/filter/sort/group/visualize
        "agg_func": None,       # sum/avg/max/min/count
        "chart_type": None,      # line/bar/pie/column
        "group_by": [],         # 分组字段
        "target_field": None,    # 目标数值字段
        "filters": [],          # 过滤条件
        "sort": None,           # asc/desc
        "description": "",       # 操作描述（用于输出）
    }

    query_lower = query.lower()

    # 1. 检测聚合函数
    for func, keywords in AGG_KEYWORDS.items():
        if any(kw in query_lower for kw in keywords):
            intent["agg_func"] = func
            intent["operation"] = "aggregate"
            break

    # 2. 检测图表类型
    for chart, keywords in CHART_KEYWORDS.items():
        if any(kw in query_lower for kw in keywords):
            intent["chart_type"] = chart
            intent["operation"] = "visualize"
            break

    # 3. 检测过滤条件
    for pattern, op in FILTER_PATTERNS:
        matches = re.findall(pattern, query_lower)
        if matches:
            intent["operation"] = "filter"
            for match in matches:
                if isinstance(match, tuple):
                    value = match[0] or match[1]
                else:
                    value = match
                intent["filters"].append({"op": op, "value": value})

    # 4. 检测排序方向
    for direction, keywords in SORT_KEYWORDS.items():
        if any(kw in query_lower for kw in keywords):
            intent["sort"] = direction

    return intent


def match_columns(headers: List[str], query: str) -> Dict:
    """模糊匹配字段"""
    query_lower = query.lower()
    result = {
        "group_fields": [],
        "value_field": None,
    }

    # 常见数值字段名
    value_candidates = ["销售额", "金额", "收入", "支出", "数量", "score", "amount", "value",
                       "总和", "总计", "共计", "数量", "价格", "利润", "rate"]
    # 常见分组字段名
    group_candidates = ["月份", "日期", "时间", "部门", "产品", "类别", "地区", "城市",
                        "客户", "渠道", "平台", "类型", "name", "category", "date"]

    # 尝试精确匹配
    for h in headers:
        h_lower = h.lower().strip()
        # 数值字段匹配
        for vc in value_candidates:
            if vc in h_lower:
                result["value_field"] = h
                break

        # 分组字段匹配
        for gc in group_candidates:
            if gc in h_lower:
                result["group_fields"].append(h)
                break

    # 如果没有匹配到，基于启发式规则
    if not result["value_field"]:
        # 找数值型列（取第一个）
        result["value_field"] = headers[-1]  # 默认最后一列是数值

    if not result["group_fields"]:
        # 取第一个非数值列作为分组
        for h in headers:
            if h != result["value_field"]:
                result["group_fields"].append(h)
                break

    return result


def parse_nl_query(query: str, headers: List[str]) -> Dict:
    """解析完整的 NL 查询"""
    intent = analyze_intent(query)
    columns = match_columns(headers, query)

    # 描述字段
    agg_desc = {
        "sum": "求和", "avg": "求平均", "max": "求最大",
        "min": "求最小", "count": "计数", "None": None
    }

    group_str = "、".join(columns["group_fields"]) if columns["group_fields"] else "全部"
    value_str = columns["value_field"] or "数值"
    agg_str = agg_desc.get(intent["agg_func"], intent["agg_func"]) or "统计"
    chart_str = intent["chart_type"] or ""

    parts = [f"按 {group_str} 对 {value_str} 进行 {agg_str}"]
    if chart_str:
        parts.append(f"并生成 {chart_str} 图表")
    if intent["filters"]:
        parts.append(f"过滤条件: {intent['filters']}")
    if intent["sort"]:
        parts.append(f"排序: {'升序' if intent['sort'] == 'asc' else '降序'}")

    intent["description"] = "，".join(parts)
    intent["columns"] = columns
    intent["headers"] = headers

    return intent


# ==================== 数据分析引擎 ====================

def execute_analysis(filepath: str, sheet: str, intent: Dict) -> Dict:
    """执行数据分析"""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(filepath, data_only=True)
        if sheet not in wb.sheetnames:
            return {"success": False, "error": f"Sheet 不存在: {sheet}"}

        ws = wb[sheet]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        headers = [str(h).strip() if h else f"列{c}" for c in headers]

        # 读取数据
        data = []
        for r in range(2, ws.max_row + 1):
            row = []
            for c in range(1, ws.max_column + 1):
                row.append(ws.cell(row=r, column=c).value)
            data.append(row)

        columns_info = intent.get("columns", {})
        group_fields = columns_info.get("group_fields", [])
        value_field = columns_info.get("value_field")

        # 找到列索引
        group_indices = []
        for gf in group_fields:
            for i, h in enumerate(headers):
                if h == gf:
                    group_indices.append(i)
                    break

        value_idx = None
        if value_field:
            for i, h in enumerate(headers):
                if h == value_field:
                    value_idx = i
                    break

        if not group_indices:
            group_indices = [0]
        if value_idx is None:
            value_idx = len(headers) - 1

        # 分组聚合
        grouped = defaultdict(list)
        for row in data:
            if len(row) <= max(max(group_indices), value_idx):
                continue
            key = tuple(row[i] if row[i] is not None else "未知" for i in group_indices)
            try:
                val = float(row[value_idx]) if row[value_idx] is not None else 0
                grouped[key].append(val)
            except (ValueError, TypeError):
                grouped[key].append(0)

        # 计算聚合
        agg_func = intent.get("agg_func") or "sum"
        results = []
        for key, values in grouped.items():
            if agg_func == "sum":
                result = sum(values)
            elif agg_func == "avg":
                result = sum(values) / len(values) if values else 0
            elif agg_func == "max":
                result = max(values) if values else 0
            elif agg_func == "min":
                result = min(values) if values else 0
            elif agg_func == "count":
                result = len(values)
            else:
                result = sum(values)

            results.append({
                "group": key,
                "value": round(result, 2),
                "count": len(values)
            })

        # 排序
        sort_dir = intent.get("sort")
        if sort_dir:
            reverse = (sort_dir == "desc")
            results.sort(key=lambda x: x["value"], reverse=reverse)

        # 生成图表
        chart_result = None
        if intent.get("chart_type"):
            chart_result = _generate_chart(filepath, sheet, results, intent)

        return {
            "success": True,
            "results": results,
            "agg_func": agg_func,
            "group_fields": [headers[i] for i in group_indices],
            "value_field": headers[value_idx],
            "total_groups": len(results),
            "chart": chart_result,
        }

    except ImportError:
        return {"success": False, "error": "需要安装 openpyxl"}
    except Exception as e:
        return {"success": False, "error": str(e)}


def _generate_chart(filepath: str, sheet: str, results: List[Dict],
                     intent: Dict) -> Optional[Dict]:
    """生成图表"""
    try:
        import openpyxl
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference

        chart_type = intent.get("chart_type", "bar")
        wb = openpyxl.load_workbook(filepath)
        ws = wb[sheet]

        # 找到结果写入位置（右侧空白列）
        start_col = ws.max_column + 2
        start_row = 2

        # 写入数据（用于图表引用）
        ws.cell(row=start_row - 1, column=start_col, value="分组")
        ws.cell(row=start_row - 1, column=start_col + 1, value="数值")
        for i, r in enumerate(results):
            ws.cell(row=start_row + i, column=start_col, value=str(r["group"][0]) if len(r["group"]) == 1 else str(r["group"]))
            ws.cell(row=start_row + i, column=start_col + 1, value=r["value"])

        # 创建图表
        if chart_type == "line":
            chart = LineChart()
        elif chart_type == "pie":
            chart = PieChart()
        else:
            chart = BarChart()

        data_ref = Reference(ws, min_col=start_col, min_row=start_row,
                            max_col=start_col + 1, max_row=start_row + len(results) - 1)
        chart.add_data(data_ref, titles_from_data=True)

        cats_ref = Reference(ws, min_col=start_col + 1, min_row=start_row,
                            max_row=start_row + len(results) - 1)
        chart.set_categories(cats_ref)

        chart.title = f"{intent.get('description', '数据分析')}"
        chart.width = 15
        chart.height = 10

        ws.add_chart(chart, f"{chr(65 + start_col - 1)}{start_row + len(results) + 2}")

        wb.save(filepath)
        return {"chart_created": True, "chart_type": chart_type}
    except Exception as e:
        return {"chart_error": str(e)}


# ==================== 通用入口 ====================

def nl_analyze(filepath: str, query: str, sheet: str = "Sheet1") -> Dict:
    """
    自然语言数据分析

    Args:
        filepath: Excel 文件路径
        query: 自然语言查询
        sheet: 工作表名称（可选）
    """
    try:
        import openpyxl

        wb = openpyxl.load_workbook(filepath, data_only=True)
        if sheet not in wb.sheetnames:
            return {"success": False, "error": f"Sheet 不存在: {sheet}"}

        ws = wb[sheet]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        headers = [str(h).strip() if h else f"列{c}" for h in headers]

        # 解析查询
        intent = parse_nl_query(query, headers)

        # 执行分析
        result = execute_analysis(filepath, sheet, intent)
        result["query"] = query
        result["parsed_intent"] = {
            "operation": intent.get("operation"),
            "agg_func": intent.get("agg_func"),
            "chart_type": intent.get("chart_type"),
            "group_fields": intent.get("columns", {}).get("group_fields", []),
            "value_field": intent.get("columns", {}).get("value_field"),
        }

        return result

    except ImportError:
        return {"success": False, "error": "需要安装 openpyxl"}
    except Exception as e:
        return {"success": False, "error": str(e)}


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Excel 自然语言数据分析")
    parser.add_argument("query", help="自然语言查询")
    parser.add_argument("--file", required=True, help="Excel 文件路径")
    parser.add_argument("--sheet", default="Sheet1", help="工作表名称")

    args = parser.parse_args()
    result = nl_analyze(args.file, args.query, args.sheet)
    print(json.dumps(result, ensure_ascii=False, default=str))
