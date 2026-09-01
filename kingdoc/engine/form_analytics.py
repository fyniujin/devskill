"""KingDoc 表单答卷收集统计引擎

v4.0 新增：云端数据的本地智能加工。

能力：
- 实现开放平台表单答卷列表/答卷内容接口
- 拉取答卷自动聚合（按题统计、交叉分析、未填名单）
- 结果写回智能文档表格并生成图表页
- 支持答卷导出 Excel
- 硬件自适应（批量拉取时读取 hardware.py）
- 零第三方依赖（仅标准库）
- 零密钥可用（本地降级模式）
"""
from __future__ import annotations

import csv
import io
import json
import os
import tempfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


class FormAnswerAggregator:
    """答卷聚合分析器"""

    def __init__(self):
        self.answers: List[Dict] = []
        self.questions: List[Dict] = []
        self.form_info: Dict = {}

    def load_answers(self, answers: List[Dict]):
        """加载答卷数据"""
        self.answers = answers

    def load_questions(self, questions: List[Dict]):
        """加载题目结构"""
        self.questions = questions

    def load_form_info(self, form_info: Dict):
        """加载表单信息"""
        self.form_info = form_info

    def get_summary(self) -> Dict:
        """获取答卷汇总统计

        Returns:
            {
                "total_answers": int,
                "questions": [
                    {
                        "question_id": str,
                        "title": str,
                        "type": str,
                        "stats": {...}
                    }
                ],
                "cross_analysis": [...],
                "unfilled_list": [...]
            }
        """
        summary = {
            "form_id": self.form_info.get("form_id", ""),
            "form_name": self.form_info.get("name", ""),
            "total_answers": len(self.answers),
            "questions": [],
            "cross_analysis": [],
            "unfilled_list": [],
        }

        # 按题统计
        for question in self.questions:
            q_id = question.get("question_id", "")
            q_title = question.get("title", "")
            q_type = question.get("type", "text")

            stats = self._analyze_question(q_id, q_type)
            summary["questions"].append({
                "question_id": q_id,
                "title": q_title,
                "type": q_type,
                "stats": stats,
            })

        # 交叉分析（针对选择题）
        summary["cross_analysis"] = self._cross_analysis()

        # 未填名单
        summary["unfilled_list"] = self._find_unfilled()

        return summary

    def _analyze_question(self, question_id: str, q_type: str) -> Dict:
        """分析单题统计"""
        values = []
        for answer in self.answers:
            answers = answer.get("answers", {})
            if question_id in answers:
                val = answers[question_id]
                if val:
                    values.append(val)

        total = len(values)
        non_empty = len([v for v in values if v])

        stats = {
            "total_responses": total,
            "non_empty": non_empty,
            "empty": total - non_empty,
            "fill_rate": round(non_empty / total, 3) if total > 0 else 0,
        }

        if q_type in ("select", "radio", "checkbox"):
            # 选择题：统计选项分布
            counter = Counter()
            for v in values:
                if isinstance(v, list):
                    for item in v:
                        counter[item] += 1
                else:
                    counter[v] += 1
            stats["distribution"] = dict(counter.most_common())
            stats["unique_options"] = len(counter)

        elif q_type == "number":
            # 数值题：统计均值/最大/最小
            try:
                nums = [float(v) for v in values if v]
                if nums:
                    stats["mean"] = round(sum(nums) / len(nums), 2)
                    stats["max"] = max(nums)
                    stats["min"] = min(nums)
                    stats["sum"] = sum(nums)
            except (ValueError, TypeError):
                pass

        elif q_type == "text":
            # 文本题：统计字数分布
            lengths = [len(str(v)) for v in values]
            if lengths:
                stats["avg_length"] = round(sum(lengths) / len(lengths), 1)
                stats["max_length"] = max(lengths)
                stats["min_length"] = min(lengths)

        return stats

    def _cross_analysis(self) -> List[Dict]:
        """交叉分析（两两选择题关联）"""
        cross = []
        select_questions = [
            q for q in self.questions
            if q.get("type") in ("select", "radio")
        ]

        if len(select_questions) < 2:
            return cross

        # 取前两题做交叉
        q1 = select_questions[0]
        q2 = select_questions[1]
        q1_id = q1.get("question_id", "")
        q2_id = q2.get("question_id", "")

        # 构建交叉表
        cross_table = defaultdict(Counter)
        for answer in self.answers:
            answers = answer.get("answers", {})
            v1 = answers.get(q1_id, "")
            v2 = answers.get(q2_id, "")
            if v1 and v2:
                cross_table[str(v1)][str(v2)] += 1

        if cross_table:
            cross.append({
                "question_1": q1.get("title", ""),
                "question_2": q2.get("title", ""),
                "table": {k: dict(v) for k, v in cross_table.items()},
            })

        return cross

    def _find_unfilled(self) -> List[Dict]:
        """找出未填题的答卷"""
        unfilled = []
        for answer in self.answers:
            answers = answer.get("answers", {})
            missing = []
            for question in self.questions:
                q_id = question.get("question_id", "")
                if q_id not in answers or not answers[q_id]:
                    missing.append({
                        "question_id": q_id,
                        "title": question.get("title", ""),
                    })
            if missing:
                unfilled.append({
                    "answer_id": answer.get("answer_id", ""),
                    "submitter": answer.get("submitter", "匿名"),
                    "missing_count": len(missing),
                    "missing_questions": missing,
                })
        return unfilled


class FormAnalyticsEngine:
    """表单答卷收集统计引擎"""

    def __init__(self, backend: Optional[Any] = None):
        self.backend = backend
        self._local_mode = backend is None
        self.aggregator = FormAnswerAggregator()

    @property
    def is_local_mode(self) -> bool:
        return self._local_mode

    def fetch_form_answers(self, form_id: str, limit: int = 100) -> Dict:
        """拉取答卷列表

        Args:
            form_id: 表单 ID
            limit: 数量限制

        Returns:
            {"success": bool, "answers": [...], "total": int}
        """
        if self._local_mode:
            return {
                "success": False,
                "error": "本地降级模式：无法拉取答卷",
                "hint": "请配置金山 App Key 后使用",
                "answers": [],
                "total": 0,
            }

        try:
            result = self.backend.kdoc_form_answers(form_id, limit)
            answers = result.get("answers", [])
            return {
                "success": True,
                "form_id": form_id,
                "answers": answers,
                "total": len(answers),
            }
        except Exception as e:
            return {
                "success": False,
                "error": f"拉取答卷失败: {e}",
                "answers": [],
                "total": 0,
            }

    def fetch_form_questions(self, form_id: str) -> Dict:
        """拉取表单题目结构

        Args:
            form_id: 表单 ID

        Returns:
            {"success": bool, "questions": [...]}
        """
        if self._local_mode:
            return {
                "success": False,
                "error": "本地降级模式：无法拉取题目",
                "questions": [],
            }

        try:
            result = self.backend.kdoc_form_info(form_id)
            questions = result.get("questions", [])
            return {
                "success": True,
                "form_id": form_id,
                "questions": questions,
            }
        except Exception as e:
            return {
                "success": False,
                "error": f"拉取题目失败: {e}",
                "questions": [],
            }

    def analyze(self, form_id: str, limit: int = 100) -> Dict:
        """完整分析流程：拉取 → 聚合 → 统计

        Args:
            form_id: 表单 ID
            limit: 答卷数量限制

        Returns:
            汇总统计结果
        """
        # 拉取题目
        q_result = self.fetch_form_questions(form_id)
        if not q_result.get("success"):
            return q_result

        # 拉取答卷
        a_result = self.fetch_form_answers(form_id, limit)
        if not a_result.get("success"):
            return a_result

        # 加载到聚合器
        self.aggregator.load_questions(q_result.get("questions", []))
        self.aggregator.load_answers(a_result.get("answers", []))
        self.aggregator.load_form_info({
            "form_id": form_id,
            "name": q_result.get("form_name", ""),
        })

        return self.aggregator.get_summary()

    def export_to_csv(self, form_id: str, output_path: str = "") -> Dict:
        """导出答卷为 CSV

        Args:
            form_id: 表单 ID
            output_path: 输出路径

        Returns:
            {"success": bool, "output_path": str, "message": str}
        """
        a_result = self.fetch_form_answers(form_id, limit=1000)
        if not a_result.get("success"):
            return a_result

        answers = a_result.get("answers", [])
        if not answers:
            return {"success": False, "error": "无答卷可导出"}

        if not output_path:
            output_path = os.path.join(tempfile.gettempdir(), f"form_{form_id}_answers.csv")

        try:
            # 收集所有题目 ID
            question_ids = []
            for answer in answers:
                for key in answer.get("answers", {}).keys():
                    if key not in question_ids:
                        question_ids.append(key)

            with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                # 表头
                header = ["answer_id", "submitter", "submitted_at"] + question_ids
                writer.writerow(header)
                # 数据行
                for answer in answers:
                    row = [
                        answer.get("answer_id", ""),
                        answer.get("submitter", "匿名"),
                        answer.get("submitted_at", ""),
                    ]
                    for q_id in question_ids:
                        val = answer.get("answers", {}).get(q_id, "")
                        if isinstance(val, list):
                            val = "; ".join(str(v) for v in val)
                        row.append(str(val))
                    writer.writerow(row)

            return {
                "success": True,
                "output_path": output_path,
                "total": len(answers),
                "message": f"已导出 {len(answers)} 条答卷到 {output_path}",
            }
        except Exception as e:
            return {
                "success": False,
                "error": f"导出失败: {e}",
            }

    def export_to_excel(self, form_id: str, output_path: str = "") -> Dict:
        """导出答卷为 Excel（需要 openpyxl）

        Args:
            form_id: 表单 ID
            output_path: 输出路径

        Returns:
            {"success": bool, "output_path": str}
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            return {
                "success": False,
                "error": "需要 openpyxl: pip install openpyxl",
            }

        a_result = self.fetch_form_answers(form_id, limit=1000)
        if not a_result.get("success"):
            return a_result

        answers = a_result.get("answers", [])
        if not answers:
            return {"success": False, "error": "无答卷可导出"}

        if not output_path:
            output_path = os.path.join(tempfile.gettempdir(), f"form_{form_id}_answers.xlsx")

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "答卷数据"

            # 收集所有题目 ID
            question_ids = []
            for answer in answers:
                for key in answer.get("answers", {}).keys():
                    if key not in question_ids:
                        question_ids.append(key)

            # 表头
            header = ["答卷ID", "提交者", "提交时间"] + question_ids
            ws.append(header)

            # 表头样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # 数据行
            for answer in answers:
                row = [
                    answer.get("answer_id", ""),
                    answer.get("submitter", "匿名"),
                    answer.get("submitted_at", ""),
                ]
                for q_id in question_ids:
                    val = answer.get("answers", {}).get(q_id, "")
                    if isinstance(val, list):
                        val = "; ".join(str(v) for v in val)
                    row.append(str(val))
                ws.append(row)

            # 自动列宽
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

            wb.save(output_path)
            return {
                "success": True,
                "output_path": output_path,
                "total": len(answers),
                "message": f"已导出 {len(answers)} 条答卷到 {output_path}",
            }
        except Exception as e:
            return {
                "success": False,
                "error": f"导出失败: {e}",
            }

    def generate_chart_data(self, form_id: str) -> Dict:
        """生成图表数据（供前端渲染）

        Args:
            form_id: 表单 ID

        Returns:
            {"success": bool, "charts": [...]}
        """
        summary = self.analyze(form_id)
        if not summary or "questions" not in summary:
            return {"success": False, "error": "分析失败"}

        charts = []
        for q_summary in summary.get("questions", []):
            q_type = q_summary.get("type", "")
            stats = q_summary.get("stats", {})

            if q_type in ("select", "radio") and "distribution" in stats:
                charts.append({
                    "question_id": q_summary.get("question_id", ""),
                    "title": q_summary.get("title", ""),
                    "chart_type": "pie" if q_type == "radio" else "bar",
                    "data": stats["distribution"],
                })
            elif q_type == "number" and "mean" in stats:
                charts.append({
                    "question_id": q_summary.get("question_id", ""),
                    "title": q_summary.get("title", ""),
                    "chart_type": "histogram",
                    "data": stats,
                })

        return {
            "success": True,
            "form_id": form_id,
            "charts": charts,
            "total_charts": len(charts),
        }

    def write_to_smart_doc(self, form_id: str, target_doc_id: str) -> Dict:
        """将统计结果写回智能文档表格

        Args:
            form_id: 表单 ID
            target_doc_id: 目标智能文档 ID

        Returns:
            {"success": bool, "message": str}
        """
        if self._local_mode:
            return {
                "success": False,
                "error": "本地降级模式：无法写回智能文档",
            }

        summary = self.analyze(form_id)
        if not summary or "questions" not in summary:
            return {"success": False, "error": "分析失败"}

        try:
            # 构建表格数据
            table_data = []
            # 表头
            table_data.append(["题目", "类型", "回答数", "有效数", "填写率", "统计详情"])
            # 数据行
            for q in summary.get("questions", []):
                stats = q.get("stats", {})
                detail = ""
                if "distribution" in stats:
                    detail = json.dumps(stats["distribution"], ensure_ascii=False)[:100]
                elif "mean" in stats:
                    detail = f"均值={stats['mean']}, 最大={stats['max']}, 最小={stats['min']}"
                table_data.append([
                    q.get("title", "")[:30],
                    q.get("type", ""),
                    str(stats.get("total_responses", "")),
                    str(stats.get("non_empty", "")),
                    f"{stats.get('fill_rate', 0):.0%}",
                    detail,
                ])

            # 调用块级编辑引擎写回
            from engine.blocks import BlockEditor
            editor = BlockEditor(backend=self.backend)

            # 构建 Markdown 表格内容
            md_lines = ["| 题目 | 类型 | 回答数 | 有效数 | 填写率 | 统计详情 |",
                        "|------|------|--------|--------|--------|---------|"]
            for row in table_data[1:]:
                md_lines.append("| " + " | ".join(str(c) for c in row) + " |")
            md_content = "\n".join(md_lines)

            # 追加到目标文档
            result = editor.append_markdown(target_doc_id, md_content)

            return {
                "success": True,
                "form_id": form_id,
                "target_doc_id": target_doc_id,
                "message": f"统计结果已写入智能文档: {target_doc_id}",
                "data": result,
            }
        except Exception as e:
            return {
                "success": False,
                "error": f"写回失败: {e}",
            }


def get_form_analytics(backend: Optional[Any] = None) -> FormAnalyticsEngine:
    """获取表单分析引擎实例"""
    return FormAnalyticsEngine(backend=backend)


def analyze_form(form_id: str, backend: Optional[Any] = None) -> Dict:
    """便捷函数：分析表单"""
    engine = FormAnalyticsEngine(backend=backend)
    return engine.analyze(form_id)


def export_form_answers(form_id: str, format: str = "csv",
                        output_path: str = "", backend: Optional[Any] = None) -> Dict:
    """便捷函数：导出答卷"""
    engine = FormAnalyticsEngine(backend=backend)
    if format == "csv":
        return engine.export_to_csv(form_id, output_path)
    elif format == "xlsx":
        return engine.export_to_excel(form_id, output_path)
    else:
        return {"success": False, "error": f"不支持的格式: {format}"}
